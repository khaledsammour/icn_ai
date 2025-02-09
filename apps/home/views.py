from django import template
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import requests
import json
from time import sleep
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver import ActionChains, Chrome
from rest_framework.views import APIView
from django.http import JsonResponse
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import undetected_chromedriver as uc
import traceback
from openpyxl import load_workbook
import io
from django.core.files.uploadedfile import InMemoryUploadedFile
from .utils import getNthChildCSSSelector, getClassBasedCSSSelector, getAbsoluteXPath, sendRequest, upload_file, get_hrefs, until_not_visible, until_visible, until_visible_click, until_visible_send_keys, until_visible_with_xpath, until_visible_xpath_click, create_browser, change_content, change_text, check_if_exist, checkImageUrl, checkProduct, click_on_overlay, correct_spelling, getImageBase64, getImageUrl, save_image, extract_top_keywords, remove_emoji, replace_dimensions, translate, unwrap_divs
import re 
from .models import Websites, Blogs, Words, Products
from airtable import Airtable
import os
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
from webdriver_manager.chrome import ChromeDriverManager
import threading
from selenium.webdriver.common.keys import Keys
from urllib.request import Request, urlopen
from sentence_transformers import SentenceTransformer, models
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import faiss
# import mysql.connector

# for w in Websites.objects.all():
#     mydb = mysql.connector.connect(
#         host="84.247.191.152",
#         user="icn-user",
#         password="icnDB#123",
#         database="ai_icn"
#     )
#     mycursor = mydb.cursor()

#     sql = "INSERT INTO home_websites (seller_id, name, base_id, table_id, require_login, email_selector, email, password_selector, password, button_selector, no_pagination, pagination_click, pagination_path, product_selector, not_contains_class, inner_selector, inside_category_selector, product_click, title_prefix, ar_title_prefix, title_prefix_selector, title_prefix_attr, title_selector, title_attr, title_suffix, title_suffix_selector, title_suffix_attr, description_selector, description_attr, key_words_selector, main_img_selector, main_img_attr, img_click, img_selector, img_attr, static_price, is_price_have_comma, price_selector, price_attr, second_price_selector, second_price_attr, is_discount, discount_selector, discount_attr, is_stuck, stuck_selector, is_feature, features_selector, features_key_selector, features_key_attr, features_value_selector, features_value_attr, en_link, ar_link, ar_selector, ar_attr, export_out_of_stuck, start_index, end_index, number_of_products, change_content, translate_english, translate_arabic) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
#     val = (w.seller_id, w.name, w.base_id, w.table_id, w.require_login, w.email_selector, w.email, w.password_selector, w.password, w.button_selector, w.no_pagination, w.pagination_click, w.pagination_path, w.product_selector, w.not_contains_class, w.inner_selector, w.inside_category_selector, w.product_click, w.title_prefix, w.ar_title_prefix, w.title_prefix_selector, w.title_prefix_attr, w.title_selector, w.title_attr, w.title_suffix, w.title_suffix_selector, w.title_suffix_attr, w.description_selector, w.description_attr, w.key_words_selector, w.main_img_selector, w.main_img_attr, w.img_click, w.img_selector, w.img_attr, w.static_price, w.is_price_have_comma, w.price_selector, w.price_attr, w.second_price_selector, w.second_price_attr, w.is_discount, w.discount_selector, w.discount_attr, w.is_stuck, w.stuck_selector, w.is_feature, w.features_selector, w.features_key_selector, w.features_key_attr, w.features_value_selector, w.features_value_attr, w.en_link, w.ar_link, w.ar_selector, w.ar_attr, w.export_out_of_stuck, w.start_index, w.end_index, w.number_of_products, w.change_content, w.translate_english, w.translate_arabic)
#     mycursor.execute(sql, val)

#     mydb.commit()

API_KEY='patKfzGeYSaMEflNh.436aae2a5ffa7285045f29714bddfcee86ae9ff624a1748533231aaede505715'
def index(request):
    context = {'segment': 'index'}

    html_template = loader.get_template('home/index.html')
    return HttpResponse(html_template.render(context, request))

def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        
        if 'media' in request.path:
            media_path = os.path.join(settings.MEDIA_ROOT, request.path.split('/media/')[-1])
            if os.path.exists(media_path):
                with open(media_path, 'rb') as media_file:
                    return HttpResponse(media_file.read(), content_type='application/octet-stream')

        context['segment'] = load_template
        context['websites'] = Websites.objects.all()

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:
        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))

class SecScrapView(APIView):
    def post(self, request, *args, **kwargs):
        options = Options()
        driver = Chrome(options=options)
        driver.maximize_window()
        url = request.data['url']
        driver.get(url)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '/page/', ".shop-container .products > .product-small", inner_selector=".image-zoom_in a")
        for href in hrefs:
            try:
                driver.get(href)
                sleep(1)
                href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                soup = BeautifulSoup(href_res, 'html.parser')
                title = soup.select_one('.product-main .product_title').get_text(strip=True)

                # Get the product price
                price_elem = soup.select_one(
                    ".product-main .price-wrapper bdi"
                )
                price = price_elem.get_text().replace('JOD', '').strip() if price_elem else ''

                # Get the main image URL
                main_image_elem = soup.select_one('.product-gallery .product-images img')
                
                image = getImageUrl(request.data['id'],main_image_elem['src']) if main_image_elem else ''

                # Get additional images
                image_elems = soup.select('.product-gallery .product-images img')
                images = [getImageUrl(request.data['id'],img['data-large_image']) for img in image_elems]

                # Check stock status
                
                in_stock = soup.select_one(".in-stock").get_text(strip=True).replace('in stock', '').strip()
                # in_stock = soup.select_one(".quantity > input.qty")['max']

                # Get product attributes content
                # description_elem = soup.select_one("#tab-description, .product-short-description")
                # product_attributes_content = correct_spelling(description_elem.get_text(strip=True), 'en-US') if description_elem else ''
                # product_attributes_content = description_elem.encode_contents() if description_elem else ''
                # until_visible_click(driver, '#tab-description')
                product_attributes_content = driver.find_element(By.CSS_SELECTOR, '#tab-description').get_attribute('innerHTML').strip() if len(driver.find_elements(By.CSS_SELECTOR, '#tab-description'))>0 else driver.find_element(By.CSS_SELECTOR, '.product-short-description').get_attribute('innerHTML').strip() if len(driver.find_elements(By.CSS_SELECTOR, '.product-short-description'))>0 else ''

                # Get keywords
                key_words_elem = soup.select_one("meta[property*='og:title']")
                
                keyWords = key_words_elem['content'] if key_words_elem else ''

                # Get discount
                discount_elem = soup.select_one(".product-main .on-sale")
                discount = discount_elem.get_text().replace('-', '').replace('%', '').strip() if discount_elem else '0'

                product_attributes_content_json = {}
                
                product_attributes = soup.select("#tab-additional_information tbody > tr")
                for attr in product_attributes:
                    if 'woocommerce-product-attributes-item--weight' not in attr['class'] and 'woocommerce-product-attributes-item--dimensions' not in attr['class']: 
                        key = attr.select_one("th:nth-child(1)").get_text(strip=True)
                        val = attr.select_one("td:nth-child(2)").get_text(strip=True)
                        product_attributes_content_json[key] = val

                driver.get(soup.select_one(".header-language-dropdown ul > li > a[hreflang*='ar']")['href'])
                sleep(1)
                ar_href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                ar_soup = BeautifulSoup(ar_href_res, 'html.parser')
                ar_title = ar_soup.select_one('.product-main .product_title').get_text(strip=True) if ar_soup.select_one('.product-main .product_title') else ''
                # ar_description_elem = ar_soup.select_one(
                #     "#tab-description, .product-short-description"
                # )
                # ar_product_attributes_content = correct_spelling(ar_description_elem.get_text(strip=True), 'ar') if ar_description_elem else ''
                # ar_product_attributes_content = ar_description_elem.encode_contents() if ar_description_elem else ''
                # until_visible_click(driver, '#tab-description, .product-short-description')
                ar_product_attributes_content = driver.find_element(By.CSS_SELECTOR, '#tab-description').get_attribute('innerHTML').strip() if len(driver.find_elements(By.CSS_SELECTOR, '#tab-description'))>0 else driver.find_element(By.CSS_SELECTOR, '.product-short-description').get_attribute('innerHTML').strip() if len(driver.find_elements(By.CSS_SELECTOR, '.product-short-description'))>0 else ''
                ar_key_words_elem = ar_soup.select_one("meta[property*='og:title']")
                ar_keyWords = ar_key_words_elem['content'] if ar_key_words_elem else ''

                ar_product_attributes_content_json = {}
                
                ar_product_attributes = ar_soup.select("#tab-additional_information tbody > tr")
                for attr in ar_product_attributes:
                    if 'woocommerce-product-attributes-item--weight' not in attr['class'] and 'woocommerce-product-attributes-item--dimensions' not in attr['class']: 
                        key = attr.select_one("th:nth-child(1)").get_text(strip=True)
                        val = attr.select_one("td:nth-child(2)").get_text(strip=True)
                        ar_product_attributes_content_json[key] = val
                # Create product dictionary
                product = {
                    "Arabic Name": ar_title,
                    "English Name": title,
                    "Arabic Description": ar_product_attributes_content if len(ar_product_attributes_content) > 3 else '',
                    "English Description": product_attributes_content if len(product_attributes_content) > 3 else '',
                    "Category Id": request.data['db_category'],
                    "Arabic Brand": "",
                    "English Brand": "",
                    "Unit Price": price,
                    "Discount Type": "Percent" if discount != "0" else "",
                    "Discount": discount if discount != "0" else "",
                    "Unit": "PC",
                    "Current Stock": in_stock,
                    "Main Image URL": image,
                    "Photos URLs": str((",").join(images)) if images else image,
                    "Video Youtube URL": "",
                    "English Meta Tags": keyWords.replace('//', ','),
                    "Arabic Meta Tags": ar_keyWords.replace('//', ','),
                    "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                    "features_ar": '' if not ar_product_attributes_content_json else json.dumps(ar_product_attributes_content_json),
                    "wholesale": "no",
                    "reference_link": href,
                }
                data.append(product)
            except Exception as e:
                print(e)
                errors.append({
                    "url": href
                })

        df = pd.DataFrame(data)
        df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)

        err_df = pd.DataFrame(errors)
        err_df.to_excel('errors.xlsx', index=False)

        driver.quit()
        return JsonResponse({})

class VikushaScrapView(APIView):
    def post(self, request, *args, **kwargs):
        driver = create_browser()
        url = request.data['url']
        driver.get(url)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '/page/', ".products > .instock a.woocommerce-LoopProduct-link")
        for href in hrefs:
            try:
                driver.get(href)
                sleep(1)
                href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                soup = BeautifulSoup(href_res, 'html.parser')
                title = soup.select_one('#main .product_title').get_text(strip=True)

                # Get the product price
                price_elem = soup.select_one(
                    ".summary .amount"
                )
                price = price_elem.get_text().replace('د.ا', '').strip() if price_elem else ''

                # Get the main image URL
                main_image_elem = soup.select_one('.zoomImg, .svi-mainsection img')
                
                image = getImageUrl(request.data['id'],main_image_elem['src']) if main_image_elem else ''

                # Get additional images
                image_elems = soup.select('.woocommerce-product-gallery__wrapper > .woocommerce-product-gallery__image > a')
                images = [getImageUrl(request.data['id'],img['href']) for img in image_elems]

                # Check stock status
                
                in_stock = soup.select_one(".quantity > input.qty")['size']

                # Get product attributes content
                description_elem = soup.select_one("#tab-description")
                product_attributes_content = description_elem.get_text(separator=' ',strip=True) if description_elem else ''

                # Get keywords
                key_words_elem = soup.select_one("meta[property*='og:title']")
                
                keyWords = key_words_elem['content'] if key_words_elem else ''

                # Get discount
                discount_elem = float(soup.select(".summary .amount")[1].get_text().replace('د.ا', '').strip()) if len(soup.select(".summary .amount"))>1 else 0
                discount = float(price) - discount_elem

                product_attributes_content_json = {}
                
                product_attributes = soup.select("#tab-additional_information tbody > tr")
                for attr in product_attributes:
                    key = attr.select_one("th:nth-child(1)").get_text(strip=True)
                    val = attr.select_one("td:nth-child(2)").get_text(strip=True)
                    product_attributes_content_json[key] = val

                # Create product dictionary
                product = {
                    "Arabic Name": translate(title),
                    "English Name": title,
                    "Arabic Description": translate(product_attributes_content) if len(product_attributes_content) > 3 else '',
                    "English Description": product_attributes_content if len(product_attributes_content) > 3 else '',
                    "Category Id": request.data['db_category'],
                    "Arabic Brand": "",
                    "English Brand": "",
                    "Unit Price": price,
                    "Discount Type": "Flat" if str(discount_elem) != "0" else "",
                    "Discount": str(discount) if str(discount_elem) != "0" else "",
                    "Unit": "PC",
                    "Current Stock": in_stock,
                    "Main Image URL": image,
                    "Photos URLs": str((",").join(images)) if images else image,
                    "Video Youtube URL": "",
                    "English Meta Tags": keyWords.replace('//', ','),
                    "Arabic Meta Tags": translate(keyWords).replace('//', ','),
                    "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                    "wholesale": "no",
                    "reference_link": href,
                }
                data.append(product)
            except Exception as e:
                print(e)
                errors.append({
                    "url": href
                })

        df = pd.DataFrame(data)
        df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)

        err_df = pd.DataFrame(errors)
        err_df.to_excel('errors.xlsx', index=False)

        driver.quit()
        return JsonResponse({})
    
class HighTechScrapView(APIView):
    def post(self, request, *args, **kwargs):
        options = Options()
        driver = Chrome(options=options)
        driver.maximize_window()
        url = request.data['url']
        driver.get(url)
        driver.execute_script("window.open('https://www.freetranslations.org/english-to-arabic-translation.html');")
        driver.switch_to.window(driver.window_handles[0])
        index = 1
        data = []
        errors = []
        hrefs = []
        driver.get(url+'?per_page=10000000')
        sleep(3)
        elements = driver.find_elements(By.CSS_SELECTOR, "aside.col-lg-9 > div > div")
        for e in elements:
            if 'Out of stock' not in e.get_attribute('innerHTML'):
                hrefs.append(e.find_element(By.CSS_SELECTOR, "a.card-img-top").get_attribute("href"))

        for href in hrefs:
            try:
                driver.get(href)
                sleep(1)
                href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                soup = BeautifulSoup(href_res, 'html.parser')
                title = soup.select_one('.page-title-overlap h1').get_text(strip=True)

                # Get the product price
                price_elem = soup.select_one(
                    ".product-details >div.h3"
                )
                price = price_elem.get_text().replace('JOD', '').strip() if price_elem else ''

                # Get the main image URL
                main_image_elem = soup.select_one('.product-gallery .image-zoom')
                
                image = getImageUrl(request.data['id'],main_image_elem['src']) if main_image_elem else ''

                # Get additional images
                image_elems = soup.select('.product-gallery .image-zoom')
                images = [getImageUrl(request.data['id'],img['src']) for img in image_elems]

                # Check stock status
                
                in_stock = '3' if soup.select_one(".product-available") else '0'

                # Get product attributes content
                description_elem = soup.select_one(".page-wrapper > div.container:nth-child(5)")
                product_attributes_content = description_elem.get_text(separator=' ',strip=True) if description_elem else ''

                # Get keywords
                key_words_elem = soup.select_one("meta[property*='og:title']")
                
                keyWords = key_words_elem['content'] if key_words_elem else ''

                # Get discount
                discount_elem = float(soup.select(".summary .amount")[1].get_text().replace('د.ا', '').strip()) if len(soup.select(".summary .amount"))>1 else 0
                discount = float(price) - discount_elem

                product_attributes_content_json = {}
                
                product_attributes = soup.select(".accordion-body > div")
                for attr in product_attributes:
                    key = attr.select_one("div:nth-child(1)").get_text(strip=True)
                    val = attr.select_one("div:nth-child(2)").get_text(strip=True)
                    product_attributes_content_json[key] = val

                # Create product dictionary
                product = {
                    "Arabic Name": translate(driver, title),
                    "English Name": title,
                    "Arabic Description": translate(driver, product_attributes_content) if len(product_attributes_content) > 3 else '',
                    "English Description": product_attributes_content if len(product_attributes_content) > 3 else '',
                    "Category Id": request.data['db_category'],
                    "Arabic Brand": "",
                    "English Brand": "",
                    "Unit Price": price,
                    "Discount Type": "Flat" if str(discount_elem) != "0" else "",
                    "Discount": str(discount) if str(discount_elem) != "0" else "",
                    "Unit": "PC",
                    "Current Stock": in_stock,
                    "Main Image URL": image,
                    "Photos URLs": str((",").join(images)) if images else image,
                    "Video Youtube URL": "",
                    "English Meta Tags": keyWords.replace('//', ','),
                    "Arabic Meta Tags": translate(driver, keyWords).replace('//', ','),
                    "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                    "wholesale": "no",
                    "reference_link": href,
                }
                data.append(product)
            except Exception as e:
                print(e)
                errors.append({
                    "url": href
                })
        index = index + 1

        df = pd.DataFrame(data)
        df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)

        err_df = pd.DataFrame(errors)
        err_df.to_excel('errors.xlsx', index=False)

        driver.quit()
        return JsonResponse({})

class GTSScrapView(APIView):
    def post(self, request, *args, **kwargs):
        translateDriver = Chrome()
        translateDriver.maximize_window()
        translateDriver.get('https://www.freetranslations.org/english-to-arabic-translation.html')
        driver = Chrome()
        driver.maximize_window()
        url = request.data['url']
        driver.get(url)
        data = []
        errors = []
        hrefs = []
        driver.get(url.replace('?dsf=stock-status-is-in-stock', '?limit=10000000&dsf=stock-status-is-in-stock'))
        sleep(3)
        elements = driver.find_elements(By.CSS_SELECTOR, ".tb_products > .product-layout")
        for e in elements:
            hrefs.append(e.find_element(By.CSS_SELECTOR, "h4 > a").get_attribute("href"))
        # driver.execute_script("window.open('https://www.freetranslations.org/english-to-arabic-translation.html');")
        # driver.switch_to.window(driver.window_handles[0])
        for href in hrefs:
            try:
                driver.get(href)
                sleep(1)
                title_selector = '.tb_system_page_title > h1'
                description_selector = '.tb_product_description'
                key_words_selector = "meta[property*='og:title']"
                product_attributes_selector = ".tb_product_attributes tbody > tr"
                href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                soup = BeautifulSoup(href_res, 'html.parser')
                title = soup.select_one(title_selector).get_text(strip=True)

                # Get the product price
                price_elem = soup.select_one(
                    "meta[property*='product:price:amount']"
                )
                price = price_elem['content'].strip() if price_elem else ''

                # Get the main image URL
                main_image_elem = soup.select_one('.tb_gallery .tb_zoom_box img')
                
                image = getImageUrl(request.data['id'], main_image_elem['src']) if main_image_elem else ''

                # Get additional images
                image_elems = soup.select('.tb_gallery .tb_listing img')
                images = [getImageUrl(request.data['id'], img['src'].replace('128x128', '1200x1200')) for img in image_elems if len(img['src'])>10]

                # Check stock status
                
                in_stock = '3' if 'In Stock' in soup.select_one("meta[property*='product:availability']")['content'] else '0'

                # Get product attributes content
                description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                # print(unwrap_divs(str(soup.select_one(description_selector).contents)))
                product_attributes_content = description_elem if description_elem and 'Previous page' not in description_elem and 'From The Manufacturer' not in description_elem.lower() else ''
                # Get keywords
                key_words_elem = soup.select_one(key_words_selector)
                
                keyWords = key_words_elem['content'] if key_words_elem else ''

                # Get discount
                discount_elem = soup.select_one(".price-savings > strong")
                discount = discount_elem.get_text().replace('JOD', '').strip() if discount_elem else '0'

                product_attributes_content_json = {}
                
                product_attributes = soup.select(product_attributes_selector)
                for attr in product_attributes:
                    key = attr.select_one("td:nth-child(1)").get_text(strip=True)
                    val = attr.select_one("td:nth-child(2)").get_text(strip=True)
                    product_attributes_content_json[key] = val

                driver.get(soup.select_one(".tb_wt_header_language_menu_system a[data-language-code*='ar']")['href'])
                sleep(1)
                ar_href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                ar_soup = BeautifulSoup(ar_href_res, 'html.parser')
                ar_title = ar_soup.select_one(title_selector).get_text(strip=True) if ar_soup.select_one(title_selector) and 'الصفحة المطلوبة لا يمكن العثور عليها' not in ar_soup.select_one(title_selector).get_text(strip=True) else title
                ar_key_words_elem = ar_soup.select_one(key_words_selector)
                ar_keyWords = ar_key_words_elem['content'] if ar_key_words_elem and 'الصفحة المطلوبة لا يمكن العثور عليها' not in ar_key_words_elem['content'] else ''

                ar_product_attributes_content_json = {}
                
                ar_product_attributes = ar_soup.select(product_attributes_selector)
                for attr in ar_product_attributes:
                    key = attr.select_one("td:nth-child(1)").get_text(strip=True)
                    val = attr.select_one("td:nth-child(2)").get_text(strip=True)
                    ar_product_attributes_content_json[key] = val
                # Create product dictionary
                product = {
                    "Arabic Name": ar_title,
                    "English Name": title,
                    "Arabic Description": translate(translateDriver, product_attributes_content) if len(product_attributes_content) > 3 else request.data['arabic_description'],
                    "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                    "Category Id": request.data['db_category'],
                    "Arabic Brand": "",
                    "English Brand": "",
                    "Unit Price": price,
                    "Discount Type": "Flat" if discount != "0" else "",
                    "Discount": discount if discount != "0" else "",
                    "Unit": "PC",
                    "Current Stock": in_stock,
                    "Main Image URL": image,
                    "Photos URLs": str((",").join(images)) if images else image,
                    "Video Youtube URL": "",
                    "English Meta Tags": keyWords.replace('//', ','),
                    "Arabic Meta Tags": ar_keyWords.replace('//', ','),
                    "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                    "features_ar": '' if not ar_product_attributes_content_json else json.dumps(ar_product_attributes_content_json),
                    "wholesale": "no",
                    "reference_link": href,
                }
                data.append(product)
            except Exception as e:
                print(e)
                errors.append({
                    "url": href
                })

        df = pd.DataFrame(data)
        df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)

        driver.quit()
        translateDriver.quit()
        return JsonResponse({})

class TXONScrapView(APIView):
    def post(self, request, *args, **kwargs):
        translateDriver = Chrome()
        translateDriver.maximize_window()
        translateDriver.get('https://www.freetranslations.org/english-to-arabic-translation.html')
        driver = Chrome()
        driver.maximize_window()
        url = request.data['url']
        driver.get(url)
        data = []
        errors = []
        hrefs = []
        index = 1
        isExist = True
        while(isExist):
            driver.get(url+'#/pageNumber=' + str(index))
            sleep(3)
            isExist = False if len(driver.find_elements(By.XPATH, "//div[contains(@class, 'ajaxFilters') and contains(@style,'display: block')]"))>0 else True
            if isExist:
                elements = driver.find_elements(By.CSS_SELECTOR, ".product-grid .product-item")
                for e in elements:
                    hrefs.append(e.find_element(By.CSS_SELECTOR, ".picture > a").get_attribute("href"))
            index = index + 1
        for href in hrefs:
            try:
                driver.get(href)
                sleep(1)
                title_selector = '.product-name'
                description_selector = '.full-description'
                key_words_selector = "meta[property*='og:title']"
                href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                soup = BeautifulSoup(href_res, 'html.parser')
                available = soup.select_one(".availability").get_text(strip=True) if soup.select_one(".availability") else 'In Stock'
                if 'in stock' in available.lower():
                    title = soup.select_one(title_selector).get_text(strip=True)

                    # Get the product price
                    old_price = soup.select_one(".old-product-price > span")
                    price_elem = soup.select_one("span[itemprop*='price']")
                    price = old_price.get_text(strip=True).replace('JOD', '').strip() if old_price else price_elem['content'].strip() if price_elem else ''
                    discount = str(float(old_price.get_text(strip=True).replace('JOD', '').strip()) - float(price_elem['content'].strip())) if old_price else '0'

                    main_image_elem = soup.select_one('img.cloudzoom')
                        
                    image = getImageUrl(request.data['id'], main_image_elem['src']) if main_image_elem else ''

                    # Get additional images
                    image_elems = soup.select('.slick-track > div > a')
                    images = [getImageUrl(request.data['id'], img['data-full-image-url']) for img in image_elems]

                    # Check stock status
                    
                    in_stock = '3'

                    # Get product attributes content
                    # description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    # print(unwrap_divs(str(soup.select_one(description_selector).contents)))
                    # product_attributes_content = description_elem if description_elem else ''
                    product_attributes_content = driver.find_element(By.CSS_SELECTOR, description_selector).get_attribute('innerHTML').strip() if len(driver.find_elements(By.CSS_SELECTOR, description_selector))>0 else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'] if key_words_elem else ''

                    # Create product dictionary
                    product = {
                        "Arabic Name": title,
                        "English Name": title,
                        "Arabic Description": translate(translateDriver, product_attributes_content) if len(product_attributes_content) > 3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Flat" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": keyWords.replace('//', ','),
                        "Arabic Meta Tags": translate(translateDriver, keyWords).replace('//', ','),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
            except Exception as e:
                print(e)
                errors.append({
                    "url": href
                })

        df = pd.DataFrame(data)
        df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)

        driver.quit()
        translateDriver.quit()
        return JsonResponse({})

class BashitiScrapView(APIView):
    def post(self, request, *args, **kwargs):
        driver = create_browser()
        url = request.data['url']
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = []
        isExist = True
        index = 1
        while (isExist):
            driver.get(url+'page/'+str(index)+'/')
            sleep(3)
            isExist = True if len(driver.find_elements(By.CSS_SELECTOR, ".products > .wd-product"))>0 else False
            if isExist:
                elements = driver.find_elements(By.CSS_SELECTOR, ".products > .wd-product")
                for e in elements:
                    if len(e.find_elements(By.CSS_SELECTOR, ".out-of-stock"))==0:
                        hrefs.append(e.find_element(By.CSS_SELECTOR, "a.product-image-link").get_attribute("href"))
            index = index + 1
        error = False
        def get_details(href):
            d = [d for d in drivers if d['working'] == False][0]
            d['working'] = True
            try:
                driver = d['driver']
                driver.get(href.replace('?lang=ar', '?lang=en'))
                title_selector = '.product_title'
                description_selector = '.woocommerce-Tabs-panel--description'
                key_words_selector = "meta[property*='og:title']"
                until_visible(driver, title_selector)
                if len(driver.find_elements(By.CSS_SELECTOR, ".summary .price .amount"))==0:
                    return
                href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                soup = BeautifulSoup(href_res, 'html.parser')
                title = soup.select_one(title_selector).get_text(strip=True)
                # Get the product price
                price_elem = soup.select_one(".summary .price .amount").get_text(strip=True)
                price = price_elem.strip().replace('JOD','').replace(',', '').strip() if price_elem else ''
                # Get discount
                discount_elem = soup.select_one(".onsale.product-label").get_text(strip=True) if soup.select_one(".onsale.product-label") else None
                discount = discount_elem.replace('-','').replace('%','').strip() if discount_elem else '0'
                # Get the main image URL
                main_image_elem = soup.select_one('.wd-carousel-wrap > .wd-carousel-item img')
                image = getImageUrl(request.data['id'], main_image_elem['src']) if main_image_elem else ''
                # Get additional images
                image_elems = soup.select('.wd-carousel-wrap > .wd-carousel-item img')
                images = [getImageUrl(request.data['id'], img['src']) for img in image_elems if len(img['src'])>10]
                # Check stock status
                in_stock = '3'
                # Get product attributes content
                description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                product_attributes_content = description_elem if description_elem else ''
                # Get keywords
                key_words_elem = soup.select_one(key_words_selector)
                keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                keywords = keyWords.split('//')
                if len(product_attributes_content)>0:
                    keywords = extract_top_keywords(product_attributes_content)
                    ar_keywords = []
                    for k in keyWords.split('//'):
                        keywords.append(k)

                    for keyW in keywords:
                        ar_keywords.append(translate(keyW))
                else:
                    ar_keywords = []
                    for keyW in keywords:
                        ar_keywords.append(translate(keyW))
                driver.get(href.replace('en/', ''))
                if len(driver.find_elements(By.CSS_SELECTOR,'.page-header h3.title'))>0:
                    ar_title = translate(title)
                    ar_product_attributes_content = translate(product_attributes_content)
                else:
                    until_visible(driver, title_selector)
                    ar_href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    ar_soup = BeautifulSoup(ar_href_res, 'html.parser')
                    ar_title = ar_soup.select_one(title_selector).get_text(strip=True)
                    ar_description_elem = ar_soup.select_one(description_selector).get_text(" ",strip=True) if ar_soup.select_one(description_selector) else ''
                    ar_product_attributes_content = ar_description_elem if ar_description_elem else ''
                product = {
                    "Arabic Name": ar_title,
                    "English Name": title,
                    "Arabic Description": ar_product_attributes_content if len(ar_product_attributes_content)>3 else request.data['arabic_description'],
                    "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                    "Category Id": request.data['db_category'],
                    "Arabic Brand": "",
                    "English Brand": "",
                    "Unit Price": price,
                    "Discount Type": "",
                    "Discount": "",
                    "Unit": "PC",
                    "Current Stock": in_stock,
                    "Main Image URL": image,
                    "Photos URLs": str((",").join(images)) if images else image,
                    "Video Youtube URL": "",
                    "English Meta Tags": ','.join(keywords),
                    "Arabic Meta Tags": ','.join(ar_keywords),
                    "features": '',
                    "features_ar": '',
                    "wholesale": "no",
                    "reference_link": href,
                }
                data.append(product)
            except Exception as e:
                error = True
                print(e)
                traceback.print_exc()
                errors.append({
                    "url": href
                })
            finally:
                d['working'] = False
        executor = ThreadPoolExecutor(max_workers=9)
        
        for href in hrefs:
            if not error:
                executor.submit(get_details, href)
        executor.shutdown(wait=True)
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
        driver.quit()
        return JsonResponse({})
    
class SportEquipmentScrapView(APIView):
    def post(self, request, *args, **kwargs):
        driver = create_browser()
        url = request.data['url']
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = []
        for i in range(5):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(1)

        elements = driver.find_elements(By.CSS_SELECTOR, ".products > .product")
        for e in elements:
            if len(e.find_elements(By.CSS_SELECTOR, ".stock.in-stock"))>0:
                hrefs.append(e.find_element(By.CSS_SELECTOR, "a.product-image-link").get_attribute("href"))

        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    title_selector = '.product_title'
                    description_selector = '#tab-description'
                    key_words_selector = "meta[property*='og:title']"
                    product_attributes_selector = ".woocommerce-product-attributes > tbody > tr"
                    until_visible(driver, title_selector)
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = translate(soup.select_one(title_selector).get_text(strip=True), dest='en')
                    # Get the product price
                    price = soup.select_one('.summary .price .amount > bdi').get_text(strip=True).replace('د.ا','').replace('.','').replace(',','.').strip() if len(soup.select(".summary .price .amount > bdi"))>1 else soup.select_one("meta[property*='product:price:amount']")['content']
                    # Get discount
                    discount_elem = soup.select_one("meta[property*='product:price:amount']")['content'] if len(soup.select(".summary .price .amount > bdi"))>1 and ' – ' not in str(soup.select_one('.summary .price .amount > bdi').parent.parent) else None
                    discount = float(price) - float(discount_elem) if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.wd-carousel-wrap > div > figure > a')
                    image = getImageUrl(request.data['id'], main_image_elem['href']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.wd-carousel-wrap > div > figure > a')
                    images = []
                    for img in image_elems:
                        if len(img['href'])>10:
                            res = getImageUrl(request.data['id'], img['href'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    product_attributes_content_json = {}                
                    product_attributes = soup.select(product_attributes_selector)
                    for attr in product_attributes:
                        key = attr.select_one("th").get_text(strip=True)
                        val = attr.select_one("td").get_text(strip=True)
                        product_attributes_content_json[translate(key, dest='en')] = translate(val, dest='en')

                    driver.get(href.replace('/en/', '/'))
                    if len(driver.find_elements(By.CSS_SELECTOR,'.page-header h3.title'))>0:
                        ar_title = translate(title)
                        ar_product_attributes_content = translate(product_attributes_content)
                    else:
                        until_visible(driver, title_selector)
                        ar_href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                        ar_soup = BeautifulSoup(ar_href_res, 'html.parser')
                        ar_title = ar_soup.select_one(title_selector).get_text(strip=True)
                        ar_description_elem = ar_soup.select_one(description_selector).get_text(" ",strip=True) if ar_soup.select_one(description_selector) else ''
                        ar_product_attributes_content = ar_description_elem if ar_description_elem else ''
                        ar_product_attributes_content_json = {}
                        ar_product_attributes = ar_soup.select(product_attributes_selector)
                        for attr in ar_product_attributes:
                            key = attr.select_one("th").get_text(strip=True)
                            val = attr.select_one("td").get_text(strip=True)
                            ar_product_attributes_content_json[key] = val
                    product = {
                        "Arabic Name": ar_title,
                        "English Name": title,
                        "Arabic Description": ar_product_attributes_content if len(ar_product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Flat" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                        "features_ar": '' if not ar_product_attributes_content_json else json.dumps(ar_product_attributes_content_json),
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
        driver.quit()
        return JsonResponse({})

class AlrefaiScrapView(APIView):
    def post(self, request, *args, **kwargs):
        driver = create_browser()
        url = request.data['url']
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '&page=', ".products-grid .product", not_contains_class='.outofstock', inner_selector='a')
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    until_visible_click(driver, 'header .open-menu')
                    if len(driver.find_elements(By.XPATH, '//*[contains(@class, "language")]/a[contains(text(),"English")]')) > 0:
                        until_visible_xpath_click(driver, '//*[contains(@class, "language")]/a[contains(text(),"English")]')
                        until_visible(driver, '.details_content .details_name')
                    title_selector = '.details_content .details_name'
                    description_selector = '.details_content .details_text'
                    key_words_selector = "meta[property*='og:title']"
                    until_visible(driver, title_selector)
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = translate(soup.select_one(title_selector).get_text(strip=True), dest='en')
                    # Get the product price
                    price = soup.select_one('.details_content .details_discount .theoldprice').get_text(strip=True).replace('JD','').replace(',','').strip() if len(soup.select(".details_content .details_discount .theoldprice"))>0 and soup.select_one('.details_content .details_discount .theoldprice').get_text(strip=True).replace('JD','').replace(',','').strip() != '' else soup.select_one('.details_content .details_price .theprice').get_text(strip=True).replace('JD','').replace(',','').strip() if len(soup.select(".details_content .details_price .theprice"))>0 else ''
                    # Get discount
                    discount_elem = soup.select_one('.details_content .details_price .theprice').get_text(strip=True).replace('%','').replace('-','').strip() if len(soup.select(".details_content .details_discount .theoldprice"))>0 and soup.select_one('.details_content .details_discount .theoldprice').get_text(strip=True).replace('JD','').replace(',','.').strip() != '' and len(soup.select(".details_content .details_price .theprice"))>0 else None
                    discount = discount_elem if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.slick-track picture img')
                    image = getImageUrl(request.data['id'], main_image_elem['src']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.slick-track picture img')
                    images = []
                    for img in image_elems:
                        if len(img['src'])>10:
                            res = getImageUrl(request.data['id'], img['src'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    if len(soup.select('.additions .addition'))>0:
                        prices = soup.select('.additions .addition')
                        for pr in prices:
                            softfix = pr.select_one('.addition-name').get_text(strip=True)
                            price = pr.select_one('.addition-price').get_text(strip=True).replace('+','').replace('JOD','').strip()
                            new_title = title + ' - ' + softfix
                            product = {
                            "Arabic Name": new_title,
                            "English Name": translate(new_title, dest='en'),
                            "Arabic Description": product_attributes_content if len(product_attributes_content)>3 else request.data['arabic_description'],
                            "English Description": translate(product_attributes_content, dest='en') if len(product_attributes_content) > 3 else request.data['description'],
                            "Category Id": request.data['db_category'],
                            "Arabic Brand": "",
                            "English Brand": "",
                            "Unit Price": price,
                            "Discount Type": "Flat" if discount != "0" else "",
                            "Discount": discount if discount != "0" else "",
                            "Unit": "PC",
                            "Current Stock": in_stock,
                            "Main Image URL": image,
                            "Photos URLs": str((",").join(images)) if images else image,
                            "Video Youtube URL": "",
                            "English Meta Tags": ','.join(keywords),
                            "Arabic Meta Tags": ','.join(ar_keywords),
                            "features": '',
                            "features_ar": '',
                            "wholesale": "no",
                            "reference_link": href,
                            }
                            data.append(product)
                    else:
                        unit = soup.select_one('.details_price .unit').get_text(strip=True) if len(soup.select('.details_price .unit'))>0 else ''
                        product = {
                            "Arabic Name": f'{title} - {unit}',
                            "English Name": translate(f'{title} - {unit}', dest='en'),
                            "Arabic Description": product_attributes_content if len(product_attributes_content)>3 else request.data['arabic_description'],
                            "English Description": translate(product_attributes_content, dest='en') if len(product_attributes_content) > 3 else request.data['description'],
                            "Category Id": request.data['db_category'],
                            "Arabic Brand": "",
                            "English Brand": "",
                            "Unit Price": price,
                            "Discount Type": "Flat" if discount != "0" else "",
                            "Discount": discount if discount != "0" else "",
                            "Unit": "PC",
                            "Current Stock": in_stock,
                            "Main Image URL": image,
                            "Photos URLs": str((",").join(images)) if images else image,
                            "Video Youtube URL": "",
                            "English Meta Tags": ','.join(keywords),
                            "Arabic Meta Tags": ','.join(ar_keywords),
                            "features": '',
                            "features_ar": '',
                            "wholesale": "no",
                            "reference_link": href,
                        }
                        data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
        driver.quit()
        return JsonResponse({})  

class DermacolScrapView(APIView):
    def post(self, request, *args, **kwargs):
        driver = create_browser()
        url = request.data['url']
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '', ".c-products__list > .c-products__item a")
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = '.h2.b-detail-desc__title'
                    key_words_selector = "meta[property*='og:title']"
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = translate(soup.select_one(title_selector).get_text(strip=True), dest='en')
                    if 'ml' not in title and len(soup.select('.b-detail-desc__unit strong'))>0:
                        title += ' ' + soup.select_one('.b-detail-desc__unit strong').get_text(strip=True)
                    # Get the main image URL
                    main_image_elem = soup.select_one('.slick-track > li > a')
                    image = getImageUrl(request.data['id'], 'https://www.dermacol.com.ar/'+main_image_elem['href']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.slick-track > li > a')
                    images = []
                    for img in image_elems:
                        if len(img['href'])>10:
                            res = getImageUrl(request.data['id'], 'https://www.dermacol.com.ar/'+img['href'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    descs = soup.select('.grid:nth-child(2) > .grid__cell > .b-content:not(.hide-mobile-up)')
                    product_attributes_content = '<div class="row">'
                    for desc in descs:
                        product_attributes_content += '<div class="col-6">'
                        product_attributes_content += '<h2 style="font-weight: bold;">' + desc.select_one('h2').get_text(strip=True) + '</h2>'
                        product_attributes_content += '<p>' + desc.get_text(strip=True).replace(desc.select_one('h2').get_text(strip=True), '') + '</p>'
                        product_attributes_content += '</div>'
                    product_attributes_content += '</div>'
                    # description_elem = soup.select_one(description_selector).get_text("\n",strip=True) if soup.select_one(description_selector) else ''
                    # product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    ar_keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        ar_keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        keywords = []
                        for k in keyWords.split('//'):
                            ar_keywords.append(k)

                        for keyW in ar_keywords:
                            keywords.append(translate(keyW, dest='en'))
                    else:
                        keywords = []
                        for keyW in keywords:
                            keywords.append(translate(keyW, dest='en'))
                    
                    product = {
                        "Arabic Name": 'Dermacol - ' + translate(title),
                        "English Name": 'Dermacol - ' + title,
                        "Arabic Description": product_attributes_content if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": translate(product_attributes_content,dest='en') if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": "0",
                        "Discount Type": "",
                        "Discount": "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)

        driver.quit()
        return JsonResponse({})  
  
class UpdateStoreScrapView(APIView):
    def post(self, request, *args, **kwargs):
        driver = create_browser()
        url = request.data['url']
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, 'page/', ".products > .wd-product a.product-image-link")
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = '.product_title'
                    key_words_selector = "meta[property*='og:title']"
                    description_selector = 'div.single-product-page > div > section:nth-child(4) .elementor-widget-wrap'
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = soup.select_one(title_selector).get_text(strip=True)
                    # Get the product price
                    price = soup.select_one('meta[property*="product:price:amount"]')['content'] if len(soup.select('meta[property*="product:price:amount"]'))>0 else ''
                    # Get discount
                    discount_elem = soup.select_one('div.single-product-page > div > section:nth-child(2) .onsale').get_text(strip=True).replace('%','').replace('-','').strip() if len(soup.select("div.single-product-page > div > section:nth-child(2) .onsale"))>0 else None
                    discount = discount_elem if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.woocommerce-product-gallery__wrapper > .wd-carousel-wrap > div > figure > a')
                    image = getImageUrl(request.data['id'], main_image_elem['href']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.woocommerce-product-gallery__wrapper > .wd-carousel-wrap > div > figure > a')
                    images = []
                    for img in image_elems:
                        if len(img['href'])>10:
                            res = getImageUrl(request.data['id'], img['href'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    product_attributes_content_json = {}                
                    ar_product_attributes_content_json = {}                
                    product_attributes = soup.select('div.single-product-page > div > section:nth-child(5) .elementor-widget-wrap table > tbody > tr')
                    for attr in product_attributes:
                        key = attr.select_one("th").get_text(strip=True)
                        val = attr.select_one("td").get_text(strip=True)
                        ar_product_attributes_content_json[translate(key)] = translate(val)
                        product_attributes_content_json[key] = val
                    product = {
                        "Arabic Name": translate(title),
                        "English Name": title,
                        "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Percent" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                        "features_ar": '' if not ar_product_attributes_content_json else json.dumps(ar_product_attributes_content_json),
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
            
        driver.quit()
        return JsonResponse({})  

class DelfyScrapView(APIView):
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        brand = request.data['brand']
        title_selector = request.data['title_selector']
        description_selector = request.data['description_selector']
        main_image_selector = request.data['main_image_selector']
        image_selector = request.data['image_selector']
        image_attr = request.data['image_attr']
        store_id = request.data['id']
        key_words_selector = "meta[property*='og:title']"
        if not file or not isinstance(file, InMemoryUploadedFile):
            return JsonResponse({"error": "No file provided or invalid file"})
        
        try:
            # Read the file and load it into openpyxl
            file_content = file.read()
            file_name = file.name
            wb = load_workbook(filename=io.BytesIO(file_content))
            sheet = wb.active
            
            # Process the Excel file (example: read cell values)
            headers = [str(cell.value).strip() for cell in sheet[1]]
            
            # Process the Excel file and format data as an array of objects
            excel_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue  # Skip empty rows
                row_data = {headers[i]: row[i] for i in range(len(headers))}
                excel_data.append(row_data)

            # You can process or save the data as needed
            # driver = create_browser()
            driver = uc.Chrome(use_subprocess=False)
            sleep(1)
            data = []
            errors = []
            error = False
            for d in excel_data:
                href = d['LINK']
                price = d['price']
                category = d['category  ID']
                stock = d['Stock'] if 'Stock' in d else '3'
                if not error:
                    try:
                        driver.get(href)
                        sleep(5)
                        if request.data['click_before_description']:
                            try:
                                if len(driver.find_elements(By.CSS_SELECTOR, '.container-main > .sticky .transition-all button'))>0:
                                    until_visible_click(driver, request.data['click_before_description'])
                            except:
                                pass
                        try:
                            until_visible(driver, title_selector)
                        except:
                            print('title_selector isnt found')
                            continue
                        driver.execute_script("""
                        var style = document.createElement('style');
                        style.type = 'text/css';
                        style.innerHTML = '.slick-track:before { content: none; } .slick-track:after { content: none; } .square-image:before { content: none; } .square-image:after { content: none; }';
                        document.getElementsByTagName('body')[0].appendChild(style);
                        """)
                        href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                        soup = BeautifulSoup(href_res, 'html.parser')
                        title = soup.select_one(title_selector).get_text(strip=True)

                        if request.data['not_in_stuck']:
                            stuck = '0' if len(soup.select(request.data['not_in_stuck']))>0 else '3'
                        else:
                            stuck = stock
                        # Get the main image URL
                        if main_image_selector:
                            main_image_elem = soup.select_one(main_image_selector)
                        else:
                            main_image_elem = soup.select_one(image_selector)
                        if store_id == '2959':
                            image = getImageBase64(driver, store_id, 'https://www.garnierarabia.com'+main_image_elem[image_attr].split('?')[0]) if main_image_elem else ''
                        else:
                            try:
                                imageSrc = 'https:' + main_image_elem[image_attr].split(',')[-1].split(' ')[1] if ',' in main_image_elem[image_attr] else  main_image_elem[image_attr]
                                image = getImageBase64(driver, store_id, imageSrc) if main_image_elem else ''
                            except Exception as e:
                                print(e)
                                pass
                        # Get additional images
                        image_elems = soup.select(image_selector)
                        images = []
                        for img in image_elems:
                            if store_id == '2959':
                                res = getImageBase64(driver, store_id, 'https://www.garnierarabia.com'+img[image_attr].split('?')[0])
                            else:
                                try:
                                    imageSrc = img[image_attr].split(',')[-1].split(' ')[0] if ',' in img[image_attr] else  img[image_attr]
                                    res = getImageBase64(driver, store_id, img[image_attr])
                                except Exception as e:
                                    print(e)
                                    pass
                            if res:
                                images.append(res)
                        description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                        product_attributes_content = description_elem if description_elem else ''
                        # Get keywords
                        key_words_elem = soup.select_one(key_words_selector)
                        keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                        keywords = keyWords.split('//')
                        if len(product_attributes_content)>0:
                            keywords = extract_top_keywords(product_attributes_content)
                            ar_keywords = []
                            for k in keyWords.split('//'):
                                keywords.append(k)

                            for keyW in keywords:
                                ar_keywords.append(translate(keyW))
                        else:
                            ar_keywords = []
                            for keyW in keywords:
                                ar_keywords.append(translate(keyW))
                        
                        product = {
                            "Arabic Name": brand + ' ' + translate(title),
                            "English Name": brand + ' ' + title,
                            "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else '',
                            "English Description": product_attributes_content if len(product_attributes_content) > 3 else '',
                            "Category Id": category,
                            "Arabic Brand": "",
                            "English Brand": "",
                            "Unit Price": price,
                            "Discount Type": "",
                            "Discount": "",
                            "Unit": "PC",
                            "Current Stock": stuck,
                            "Main Image URL": image,
                            "Photos URLs": str((",").join(images)) if images else image,
                            "Video Youtube URL": "",
                            "English Meta Tags": ','.join(keywords),
                            "Arabic Meta Tags": ','.join(ar_keywords),
                            "features": '',
                            "features_ar": '',
                            "wholesale": "no",
                            "reference_link": href,
                        }
                        data.append(product)
                    except Exception as e:
                        error = True
                        print(e)
                        traceback.print_exc()
                        errors.append({
                            "url": href
                        })   
            if len(errors)>0:
                err_df = pd.DataFrame(errors)
                err_df.to_excel('excel/'+file_name+'_errors.xlsx', index=False)
            else:
                df = pd.DataFrame(data)
                df.to_excel('excel/'+file_name+'_products.xlsx', index=False)
                change_content(driver, data, request.data['db_category'])
                
            driver.quit()
            return JsonResponse({})  
        except Exception as e:
            return JsonResponse({"error": str(e)})
  
class RealCosmeticsScrapView(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '&page=', ".tab-content .row > div.col-product .product-img > a")
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = '.product-details-content > h1'
                    key_words_selector = "meta[property*='og:title']"
                    # description_selector = '.product-block-list__item--description .card__section'
                    try:
                        until_visible(driver, '.product-details > img')
                    except: 
                        pass
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = soup.select_one(title_selector).get_text(strip=True)
                    # Get the product price
                    price = soup.select_one('.product-details-price span.old').get_text(strip=True).replace('JOD','').replace(',','').strip() if len(soup.select('.product-details-price span.old'))>0 else soup.select_one('.product-details-price span').get_text(strip=True).replace('JOD','').replace(',','').strip()
                    # Get discount
                    discount_elem = soup.select_one('.product-details > span').get_text(strip=True).replace('%','').strip() if len(soup.select(".product-details > span"))>0 and '%' in soup.select_one('.product-details > span').get_text(strip=True) else None
                    discount = discount_elem if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.product-details > img')
                    image = getImageBase64(driver, request.data['id'], main_image_elem['src']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.product-details > img')
                    images = []
                    for img in image_elems:
                        if len(img['src'])>10:
                            res = getImageBase64(driver, request.data['id'], img['src'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    # description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    # product_attributes_content = description_elem if description_elem else ''
                    product_attributes_content = ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    product = {
                        "Arabic Name": translate(title),
                        "English Name": title,
                        "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Percent" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
            
        driver.quit()
        return JsonResponse({})  

class NewVisionScrapView(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, 'page/', ".wd-products > .wd-product a.product-image-link")
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = '.product_title'
                    key_words_selector = "meta[property*='og:title']"
                    description_selector = '.woocommerce-product-details__short-description'
                    try:
                        until_visible(driver, '.product-image-summary-inner .woocommerce-product-gallery__wrapper .wd-carousel-wrap figure > a')
                    except: 
                        pass
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = soup.select_one(title_selector).get_text(strip=True)
                    # Get the product price
                    price = soup.select_one('.single-product-page .price.pewc-main-price .woocommerce-Price-amount.amount').get_text(strip=True).replace('JOD','').replace(',','').strip() if len(soup.select('.single-product-page .price.pewc-main-price .woocommerce-Price-amount.amount'))>0 else ''
                    # Get discount
                    discount_elem = soup.select('.single-product-page .price.pewc-main-price .woocommerce-Price-amount.amount')[1].get_text(strip=True).replace('JOD','').replace(',','').strip() if len(soup.select(".single-product-page .price.pewc-main-price .woocommerce-Price-amount.amount"))>1 else None
                    discount = float(price) - float(discount_elem) if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.product-image-summary-inner .woocommerce-product-gallery__wrapper .wd-carousel-wrap figure > a')
                    image = getImageBase64(driver, request.data['id'], main_image_elem['href']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.product-image-summary-inner .woocommerce-product-gallery__wrapper .wd-carousel-wrap figure > a')
                    images = []
                    for img in image_elems:
                        if len(img['href'])>10:
                            res = getImageBase64(driver, request.data['id'], img['href'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    product = {
                        "Arabic Name": translate(title),
                        "English Name": title,
                        "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Flat" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
            
        driver.quit()
        return JsonResponse({})  

class BashitiCentralScrapView(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '&page=', ".all-prodict-item-list .single-item", not_contains_class='.ribbon-corner', inner_selector='.product-content-2 > h2 > a')
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = '.single-product-content .product-title:nth-child(2)'
                    key_words_selector = "meta[property*='og:title']"
                    description_selector = '.single-product-content .product-title:nth-child(5)'
                    try:
                        until_visible(driver, '.easyzoom > a')
                    except: 
                        pass
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = soup.select_one(title_selector).get_text(strip=True)
                    # Get the product price
                    price = soup.select_one('.single-product-content > .single-product-price').get_text(strip=True).replace('JOD','').replace(',','').strip() if len(soup.select('.single-product-content > .single-product-price'))>0 else ''
                    # Get discount
                    discount_elem = soup.select_one('.wd-gallery-images .onsale').get_text(strip=True).replace('-','').replace('%','').strip() if len(soup.select(".wd-gallery-images .onsale"))>0 else None
                    discount = discount_elem if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.easyzoom > a')
                    image = getImageUrl(request.data['id'], main_image_elem['href']) if main_image_elem else ''
                    print(image)
                    # Get additional images
                    image_elems = soup.select('.easyzoom > a')
                    images = []
                    for img in image_elems:
                        if len(img['href'])>10:
                            res = getImageUrl(request.data['id'], img['href'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))

                    product_attributes_content_json = {}                
                    ar_product_attributes_content_json = {}                
                    product_attributes = soup.select('#tab-additional_information tbody > tr')
                    for attr in product_attributes:
                        key = attr.select_one("th").get_text(strip=True)
                        val = attr.select_one("td").get_text(strip=True)
                        product_attributes_content_json[key] = val
                        ar_product_attributes_content_json[translate(key)] = translate(val)

                    product = {
                        "Arabic Name": translate(title),
                        "English Name": title,
                        "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Percent" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                        "features_ar": '' if not ar_product_attributes_content_json else json.dumps(ar_product_attributes_content_json),
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
            
        driver.quit()
        return JsonResponse({})  

class PetsJoScrapView(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '#/pageSize=150&viewMode=grid&orderBy=0&pageNumber=', ".product-grid > .item-grid .picture > a", should_not_exist='#nopAjaxFiltersNoProductsDialog_wnd_title')
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = '.overview .product-name'
                    key_words_selector = "meta[property*='og:title']"
                    description_selector = '.overview .short-description'
                    try:
                        until_visible(driver, '.gallery .picture > a > img')
                    except: 
                        pass
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = soup.select_one(title_selector).get_text(strip=True)
                    # Get the product price
                    price = soup.select_one('.prices > .product-price').get_text(strip=True).replace('د.ا.','').replace(',','').strip() if len(soup.select('.prices > .product-price'))>0 else soup.select_one('.product-single-details .price-box .new-price').get_text(strip=True).replace('JD','').replace(',','').strip() if len(soup.select('.product-single-details .price-box .new-price'))>0 else ''
                    # Get discount
                    discount_elem = soup.select_one('.product-single-details .price-box .new-price').get_text(strip=True).replace('JD','').replace(',','').strip() if len(soup.select(".product-single-details .price-box .old-price"))>0 else None
                    discount = float(price) - float(discount_elem) if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.gallery .picture > a > img')
                    image = getImageBase64(driver, request.data['id'], main_image_elem['src']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.gallery .picture > a > img')
                    images = []
                    for img in image_elems:
                        if len(img['src'])>10:
                            res = getImageBase64(driver, request.data['id'], img['src'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    product = {
                        "Arabic Name": translate(title),
                        "English Name": title,
                        "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Flat" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
            
        driver.quit()
        return JsonResponse({})  
  
class PetsJoScrapView(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '?page=', ".product-grid > .item-grid .picture > a", should_not_exist='#nopAjaxFiltersNoProductsDialog_wnd_title')
        error = False
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = '.overview .product-name'
                    key_words_selector = "meta[property*='og:title']"
                    description_selector = '.overview .short-description'
                    try:
                        until_visible(driver, '.gallery .picture > a > img')
                    except: 
                        pass
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = soup.select_one(title_selector).get_text(strip=True)
                    # Get the product price
                    price = soup.select_one('.prices > .product-price').get_text(strip=True).replace('د.ا.','').replace(',','').strip() if len(soup.select('.prices > .product-price'))>0 else soup.select_one('.product-single-details .price-box .new-price').get_text(strip=True).replace('JD','').replace(',','').strip() if len(soup.select('.product-single-details .price-box .new-price'))>0 else ''
                    # Get discount
                    discount_elem = soup.select_one('.product-single-details .price-box .new-price').get_text(strip=True).replace('JD','').replace(',','').strip() if len(soup.select(".product-single-details .price-box .old-price"))>0 else None
                    discount = float(price) - float(discount_elem) if discount_elem else '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.gallery .picture > a > img')
                    image = getImageBase64(driver, request.data['id'], main_image_elem['src']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.gallery .picture > a > img')
                    images = []
                    for img in image_elems:
                        if len(img['src'])>10:
                            res = getImageBase64(driver, request.data['id'], img['src'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    product = {
                        "Arabic Name": translate(title),
                        "English Name": title,
                        "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Flat" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
            
        driver.quit()
        return JsonResponse({})  
 
class ArabiEmartScrapView(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        data = []
        errors = []
        hrefs = get_hrefs(driver, url, '&page=', ".grid .card header a.no-underline", index=int(request.data['index']), max_index=int(request.data['max_index']), start_pagination=True)
        error = False
        for i, e in enumerate(hrefs):
            match = re.search(r'"item_url":"(/items/en/[^"]+)"', e)
            if match:
                item_url = match.group(1)
                hrefs[i] = 'https://arabiemart.com' + item_url
            else:
                print("item_url not found")
                
        for href in hrefs:
            if not error:
                try:
                    driver.get(href)
                    sleep(1)
                    title_selector = 'article .bigtitle'
                    key_words_selector = "meta[property*='og:title']"
                    description_selector = 'article .whitespace-pre-wrap'
                    try:
                        until_visible(driver, '.relative.bg-panel picture img')
                    except: 
                        pass
                    if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                        continue
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    title = soup.select_one(title_selector).get_text(strip=True)
                    # Get the product price
                    price = soup.select_one('article small.sidenote').get_text().replace('JOD','').replace(',','').strip() if len(soup.select('article small.sidenote'))>0 else soup.select_one('article strong[data="item-price"]').get_text(strip=True).replace('JOD','').replace(',','').strip() if len(soup.select('article strong[data="item-price"]'))>0 else ''
                    # Get discount
                    discount = '0'
                    # Get the main image URL
                    main_image_elem = soup.select_one('.relative.bg-panel picture img')
                    image = getImageBase64(driver, request.data['id'], main_image_elem['src']) if main_image_elem else ''
                    # Get additional images
                    image_elems = soup.select('.relative.bg-panel picture img')
                    images = []
                    for img in image_elems:
                        if len(img['src'])>10:
                            res = getImageBase64(driver, request.data['id'], img['src'])
                            if res:
                                images.append(res)
                    # Check stock status
                    in_stock = '3'
                    # Get product attributes content
                    description_elem = soup.select_one(description_selector).get_text(" ",strip=True) if soup.select_one(description_selector) else ''
                    product_attributes_content = description_elem if description_elem else ''
                    # Get keywords
                    key_words_elem = soup.select_one(key_words_selector)
                    keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                    keywords = keyWords.split('//')
                    if len(product_attributes_content)>0:
                        keywords = extract_top_keywords(product_attributes_content)
                        ar_keywords = []
                        for k in keyWords.split('//'):
                            keywords.append(k)

                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    else:
                        ar_keywords = []
                        for keyW in keywords:
                            ar_keywords.append(translate(keyW))
                    
                    product = {
                        "Arabic Name": translate(title),
                        "English Name": title,
                        "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                        "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                        "Category Id": request.data['db_category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": price,
                        "Discount Type": "Flat" if discount != "0" else "",
                        "Discount": discount if discount != "0" else "",
                        "Unit": "PC",
                        "Current Stock": in_stock,
                        "Main Image URL": image,
                        "Photos URLs": str((",").join(images)) if images else image,
                        "Video Youtube URL": "",
                        "English Meta Tags": ','.join(keywords),
                        "Arabic Meta Tags": ','.join(ar_keywords),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": href,
                    }
                    data.append(product)
                except Exception as e:
                    error = True
                    print(e)
                    traceback.print_exc()
                    errors.append({
                        "url": href
                    })   
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['max_index']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['max_index']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['max_index'])
            
        driver.quit()
        return JsonResponse({})  
 
mainExecutor = None
class MainScrapView(APIView):
    def post(self, request, *args, **kwargs):
        website = Websites.objects.get(name=request.data['name'])
        baseId = website.base_id
        tableId = website.table_id
        viewId = request.data['viewId']
        urlKey = 'Link'
        categoryKey = 'Category'
        airtable = Airtable(baseId, API_KEY)
        all_records = airtable.iterate(tableId, view=viewId)
        def process_record(r):
            url = r['fields'][urlKey]
            category = r['fields'][categoryKey]
            prefix_arabic = r['fields'].get('Arabic Name', '')
            prefix_english= r['fields'].get('English Name', '')
            driver = create_browser(page_load_strategy='eager' if request.data['name'] == 'Indola stores' else 'normal')
            driver.get(url)
            sleep(3)

            if len(driver.find_elements(By.CSS_SELECTOR,'.flatsome-cookies .button.primary'))>0:
                until_visible_click(driver, '.flatsome-cookies .button.primary')
                
            if len(driver.find_elements(By.CSS_SELECTOR,'.popup[style="display: block !important;visibility: visible !important;"] .close'))>0:
                driver.find_element(By.CSS_SELECTOR,'.popup[style="display: block !important;visibility: visible !important;"] .close').click()
            data = []
            errors = []
            error = False
            def product_details(href):
                try:
                    until_visible(driver, website.title_selector)
                except: 
                    pass
                if len(driver.find_elements(By.CSS_SELECTOR, website.title_selector))==0:
                    return
                if website.no_detail_page:
                    soup = BeautifulSoup(href.get_attribute('innerHTML'), 'html.parser')
                else:
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                if website.title_attr:
                    title = soup.select_one(website.title_selector)[website.title_attr].strip()
                else:
                    title = soup.select_one(website.title_selector).get_text(strip=True)

                title_prefix = ''
                if website.title_prefix:
                    title_prefix = website.title_prefix
                    title = title.replace(title_prefix,'').strip()

                if prefix_english:
                    title_prefix = prefix_english
                    title = title.replace(prefix_english,'').strip()

                if website.title_prefix_attr:
                    if len(soup.select(website.title_prefix_selector))>0:
                        title_prefix = soup.select_one(website.title_prefix_selector)[website.title_prefix_attr].strip()
                elif website.title_prefix_selector:
                    if len(soup.select(website.title_prefix_selector))>0:
                        title_prefix = soup.select_one(website.title_prefix_selector).get_text(strip=True)

                ar_title_prefix = ''
                if website.ar_title_prefix:
                    ar_title_prefix = website.ar_title_prefix

                if prefix_arabic:
                    ar_title_prefix = prefix_arabic
                    
                title_suffix = ''
                if website.title_suffix:
                    title_suffix = website.title_suffix
                    
                if website.title_suffix_attr:
                    if len(soup.select(website.title_suffix_selector))>0:
                        title_suffix = soup.select_one(website.title_suffix_selector)[website.title_suffix_attr].strip()
                elif website.title_suffix_selector:
                    if len(soup.select(website.title_suffix_selector))>0:
                        title_suffix = soup.select_one(website.title_suffix_selector).get_text(strip=True)
                # Get the product price
                price = ''
                if website.static_price:
                    price = website.static_price
                elif website.price_attr:
                    price = soup.select_one(website.price_selector)[website.price_attr].replace('Regular price','').replace('.أ.د','').replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').replace(',','').strip() if len(soup.select(website.price_selector))>0 else ''
                else:
                    if website.price_selector and len(soup.select(website.price_selector))>0:
                        for prices in soup.select(website.price_selector):
                            try:
                                if float(prices.get_text(strip=True).replace('Regular price','').replace('.أ.د','').replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').strip().replace(',', '')):
                                    price = prices.get_text(strip=True).replace('Regular price','').replace('.أ.د','').replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').strip() if prices.get_text(strip=True).replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').strip() != '0.000' else ''
                            except Exception as e:
                                print('price error: ', e)
                                pass

                    if website.second_price_attr and not price:
                        price = soup.select_one(website.second_price_selector)[website.second_price_attr].replace('Regular price','').replace('.أ.د','').replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').replace(',','').strip() if len(soup.select(website.second_price_selector))>0 else ''
                    if len(price)==0 and website.second_price_selector and len(soup.select(website.second_price_selector))>0:
                        price = soup.select_one(website.second_price_selector).get_text(strip=True).replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').strip() if website.second_price_selector and len(soup.select(website.second_price_selector))>0 else ''
                    if website.is_price_have_comma:
                        price = price.replace(',','.')

                discount = '0'
                discount_elem = None
                if website.discount_attr:
                    discount = soup.select_one(website.discount_selector)[website.discount_attr].replace('Regular price','').replace('.أ.د','').replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').replace(',','').strip() if len(soup.select(website.discount_selector))>0 else ''
                else:
                    if website.discount_selector and len(soup.select(website.discount_selector))>0:
                        discount_elem = soup.select_one(website.discount_selector).get_text(strip=True).replace('Regular price','').replace('.أ.د','').replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').strip() if len(soup.select(website.discount_selector))>0 and soup.select_one(website.discount_selector).get_text(strip=True).replace('د.ا', '').replace('JD','').replace('JOD','').replace('د.أ', '').replace('السعر/قطعة', '').strip() != '0.000' else ''
                        discount = price
                        price = discount_elem
                        discount = float(price) - float(discount) if discount else '0'
                # Get the main image URL
                main_image_elem = soup.select_one(website.main_img_selector)
                if main_image_elem:
                    if website.main_img_attr == 'style':
                        style = main_image_elem.get('style', '')
                        url_match = re.search(r'background-image:\s*url\(["\']?(.*?)["\']?\)', style)
                        image_url = url_match.group(1) if url_match else ''
                        image = getImageBase64(driver, website.seller_id, image_url) if image_url else ''
                    else:
                        image = getImageBase64(driver, website.seller_id, main_image_elem[website.main_img_attr])
                else:
                    image = ''
                # Get additional images
                images = []
                if website.img_click:
                    image_elems = driver.find_elements(By.CSS_SELECTOR, website.img_selector)
                    for indx, i in enumerate(image_elems):
                        try:
                            until_visible_click(driver, website.img_selector + ':nth-child('+str(indx+1)+')')
                            sleep(2)
                            until_visible(driver, website.main_img_selector)
                            img = driver.find_element(By.CSS_SELECTOR, website.main_img_selector)
                            if website.main_img_attr == 'style':
                                print(img)
                                style = img.get_attribute('style')
                                url_match = re.search(r'background-image:\s*url\(["\']?(.*?)["\']?\)', style)
                                image_url = url_match.group(1) if url_match else ''
                                image = getImageBase64(driver, website.seller_id, image_url) if image_url else ''
                            elif len(img.get_attribute(website.main_img_attr))>10:
                                res = getImageBase64(driver, website.seller_id, img.get_attribute(website.main_img_attr))
                                if res:
                                    images.append(res)
                        except Exception as e:
                            print(e)
                            pass
                else:
                    image_elems = soup.select(website.img_selector)
                    for img in image_elems:
                        if website.img_attr == 'style':
                            style = img.get_attribute('style')
                            url_match = re.search(r'background-image:\s*url\(["\']?(.*?)["\']?\)', style)
                            image_url = url_match.group(1) if url_match else ''
                            image = getImageBase64(driver, website.seller_id, image_url) if image_url else ''
                        elif len(img[website.img_attr])>10:
                            res = getImageBase64(driver, website.seller_id, img[website.img_attr])
                            if res:
                                images.append(res)
                # Check stock status
                if website.is_stuck:
                    in_stock = '100' if len(soup.select(website.stuck_selector))==0 else '0'
                else:
                    in_stock = '100'
                # Get product attributes content
                if website.description_selector:
                    if website.description_attr:
                        description_elem = soup.select_one(website.description_selector)[website.description_attr] if soup.select_one(website.description_selector) else ''
                    else:
                        description_elem = soup.select_one(website.description_selector).get_text(" ",strip=True) if soup.select_one(website.description_selector) else ''
                else:
                    description_elem = ''
                product_attributes_content = description_elem if description_elem else ''
                # Get keywords
                key_words_elem = soup.select_one(website.key_words_selector)
                keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                keywords = keyWords.split('//')
                if len(product_attributes_content)>0:
                    keywords = extract_top_keywords(product_attributes_content)
                    ar_keywords = []
                    for k in keyWords.split('//'):
                        keywords.append(translate(k, dest="en"))

                    for keyW in keywords:
                        ar_keywords.append(translate(keyW))
                else:
                    ar_keywords = []
                    for keyW in keywords:
                        ar_keywords.append(translate(keyW))

                if website.en_link or website.ar_selector:
                    product_attributes_content_json = {}                
                    if website.is_feature:            
                        product_attributes = soup.select(website.features_selector)
                        for attr in product_attributes:
                            if attr.select_one(website.features_key_selector) and attr.select_one(website.features_value_selector):
                                key = attr.select_one(website.features_key_selector).get_text(strip=True)
                                val = attr.select_one(website.features_value_selector).get_text(strip=True)
                                product_attributes_content_json[translate(key, dest="en")] = translate(val, dest="en")
                    if website.ar_selector:
                        ar_href =  soup.select_one(website.ar_selector)
                        driver.get(ar_href[website.ar_attr])
                    elif website.ar_link:
                        driver.get(driver.current_url.replace(website.en_link, website.ar_link))
                    else:
                        driver.get(driver.current_url.replace(website.en_link, ''))
                    
                    try:
                        until_visible(driver, website.title_selector)
                    except: 
                        pass
                    if len(driver.find_elements(By.CSS_SELECTOR, website.title_selector))==0:
                        return
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    if website.title_attr:
                        ar_title = soup.select_one(website.title_selector)[website.title_attr].strip()
                    else:
                        ar_title = soup.select_one(website.title_selector).get_text(strip=True)

                    if website.description_selector:
                        if website.description_attr:
                            ar_description_elem = soup.select_one(website.description_selector)[website.description_attr] if soup.select_one(website.description_selector) else ''
                        else:
                            ar_description_elem = soup.select_one(website.description_selector).get_text(" ",strip=True) if soup.select_one(website.description_selector) else ''
                    else:
                        ar_description_elem = None

                    ar_product_attributes_content = ar_description_elem if ar_description_elem else ''
                    ar_description = ar_product_attributes_content if len(ar_product_attributes_content)>3 else request.data['arabic_description']
                    ar_product_attributes_content_json = {}    
                    if website.is_feature:            
                        product_attributes = soup.select(website.features_selector)
                        for attr in product_attributes:
                            if attr.select_one(website.features_key_selector) and attr.select_one(website.features_value_selector):
                                key = attr.select_one(website.features_key_selector).get_text(strip=True)
                                val = attr.select_one(website.features_value_selector).get_text(strip=True)
                                ar_product_attributes_content_json[translate(key)] = translate(val)
                else:
                    if website.translate_arabic:
                        ar_title = translate(title, source='en')
                    else:
                        ar_title = title
                    if ar_title_prefix:
                        ar_title = ar_title.replace(ar_title_prefix,'').strip()
                    # ar_title = title
                    ar_description = translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description']
                    product_attributes_content_json = {}                
                    ar_product_attributes_content_json = {}    
                    if website.is_feature:            
                        product_attributes = soup.select(website.features_selector)
                        for attr in product_attributes:
                            if attr.select_one(website.features_key_selector) and attr.select_one(website.features_value_selector):
                                key = attr.select_one(website.features_key_selector).get_text(strip=True)
                                val = attr.select_one(website.features_value_selector).get_text(strip=True)
                                ar_product_attributes_content_json[translate(key)] = translate(val)
                                product_attributes_content_json[translate(key, dest="en")] = translate(val, dest="en")
                product = {
                    "Arabic Name": (ar_title_prefix+' ' if ar_title_prefix else translate(title_prefix, source='en') +' ' if title_prefix else '') + ar_title +  (' ' + translate(title_suffix) if title_suffix else ''),
                    "English Name": (title_prefix +' ' if title_prefix else '') + (translate(title, dest="en") if website.translate_english else title) +  (' ' + title_suffix if title_suffix else ''),
                    "Arabic Description": ar_description.replace('الوصف','').strip(),
                    "English Description": translate(product_attributes_content.replace('Description', '').strip(), dest="en") if len(product_attributes_content) > 3 else request.data['description'],
                    "Category Id": category,
                    "Arabic Brand": "",
                    "English Brand": "",
                    "Unit Price": price,
                    "Discount Type": "Flat" if discount != "0" else "",
                    "Discount": discount if discount != "0" else "",
                    "Unit": "PC",
                    "Current Stock": in_stock,
                    "Main Image URL": image,
                    "Photos URLs": str((",").join(images)) if images else image,
                    "Video Youtube URL": "",
                    "English Meta Tags": ','.join(keywords),
                    "Arabic Meta Tags": ','.join(ar_keywords),
                    "features": '' if not product_attributes_content_json else json.dumps(product_attributes_content_json),
                    "features_ar": '' if not ar_product_attributes_content_json else json.dumps(ar_product_attributes_content_json),
                    "wholesale": "no",
                    "reference_link": href,
                }
                data.append(product)
        
            if website.require_login:
                if len(driver.find_elements(By.CSS_SELECTOR, website.email_selector))>0:
                    until_visible_send_keys(driver, website.email_selector, website.email)
                    until_visible_send_keys(driver, website.password_selector, website.password)
                    until_visible_click(driver, website.button_selector)
                    sleep(3)
            # if len(driver.find_elements(By.CSS_SELECTOR, website.main_img_selector))>0:
            #     product_details(url)
            elif website.product_click:
                isExist = True
                index = website.start_index
                fitst_index = index
                pagination = website.pagination_path
                while(isExist):
                    if index != fitst_index:
                        driver.get(url+pagination+str(index)+('/' if 'page=' not in pagination and 'pageNumber=' not in pagination and pagination != 'p=' else ''))
                    sleep(3)
                    isExist = check_if_exist(driver, website.product_selector, "products")
                    elements = driver.find_elements(By.CSS_SELECTOR,  website.product_selector)
                    u = driver.current_url
                    for i, e in enumerate(elements):
                        driver.get(u)
                        until_visible_click(driver, f'{website.product_selector}:nth-child({i+1})')
                        if website.number_of_products and website.number_of_products == i:
                            break
                        if not error:
                            try:
                                product_details(driver.current_url)
                            except Exception as e:
                                error = True
                                print(e)
                                traceback.print_exc()
                                errors.append({
                                    "url": driver.current_url
                                })
                    if website.no_pagination:
                        isExist = False
            elif website.no_detail_page:
                for d in driver.find_elements(By.CSS_SELECTOR, 'div[data-testid="accordion"]'):
                    d.click()
                while(True):
                    if len(driver.find_elements(By.CSS_SELECTOR, website.pagination_click))==0:
                        break
                    try:
                        until_visible_click(driver, website.pagination_click)
                    except:
                        pass

                hrefs = driver.find_elements(By.CSS_SELECTOR, website.product_selector)
                for href in hrefs:
                    if not error:
                        try:
                            product_details(href)
                        except Exception as e:
                            error = True
                            print(e)
                            traceback.print_exc()
                            errors.append({
                                "url": href
                            })
            else:
                hrefs = get_hrefs(driver, url, website.pagination_path, website.product_selector, index=website.start_index, no_pagination=website.no_pagination, pagination_click=website.pagination_click, not_contains_class=website.not_contains_class, inner_selector=website.inner_selector)
                if website.inside_category_selector:
                    driver.get(url)
                    category_hrefs = get_hrefs(driver, url, '/', website.inside_category_selector, index=1, no_pagination=True)
                    for cHref in category_hrefs:
                        driver.get(cHref)
                        newHref = get_hrefs(driver, cHref, website.pagination_path, website.product_selector, index=website.start_index, no_pagination=website.no_pagination)
                        for newH in newHref:
                            hrefs.append(newH)
                if website.number_of_products:
                    hrefs = hrefs[:website.number_of_products]

                for href in hrefs:
                    if not error:
                        try:
                            driver.get(href)
                            sleep(1)
                            product_details(href)
                        except Exception as e:
                            error = True
                            print(e)
                            traceback.print_exc()
                            errors.append({
                                "url": href
                            })   
            if len(errors)>0:
                err_df = pd.DataFrame(errors)
                err_df.to_excel('excel/'+r['id']+'_errors.xlsx', index=False)
            else:
                df = pd.DataFrame([d for d in data if d['Current Stock'] != '0'])
                df.to_excel('excel/'+r['id']+'_products.xlsx', index=False)
                if website.change_content:
                    change_content(driver, [d for d in data if d['Current Stock'] != '0'], r['id'])
                else:
                    try:
                        url = "https://ai.icn.com/api/upload_image"
                        files = {
                            'file': (r['id']+'_products.xlsx', open(os.path.join('excel', r['id']+'_products.xlsx'), 'rb'))  # Open the image in binary mode
                        }
                        data = {
                            'base_id': baseId,
                            'table_id': tableId,
                            'record_id': r['id'],
                        }
                        # Send the POST request
                        response = requests.post(url, files=files, data=data)
                        return response.json()  # Return the response as JSON
                    except Exception as error:
                        print(error)
                if website.export_out_of_stuck:
                    df = pd.DataFrame([d for d in data if d['Current Stock'] == '0'])
                    df.to_excel('excel/'+r['id']+'out_products.xlsx', index=False)
                    change_content(driver, [d for d in data if d['Current Stock'] == '0'], r['id']+'out', withoutReset=False)
                if os.path.join('excel', 'new_'+r['id']+'_products.xlsx'):
                    try:
                        url = "https://ai.icn.com/api/upload_image"
                        files = {
                            'file': ('new_'+r['id']+'_products.xlsx', open(os.path.join('excel', 'new_'+r['id']+'_products.xlsx'), 'rb'))  # Open the image in binary mode
                        }
                        data = {
                            'base_id': baseId,
                            'table_id': tableId,
                            'record_id': r['id'],
                        }
                        # Send the POST request
                        response = requests.post(url, files=files, data=data)
                        return response.json()  # Return the response as JSON
                    except Exception as error:
                        print(error)
            driver.quit()
        with ThreadPoolExecutor(max_workers=6) as executor:
            global mainExecutor
            mainExecutor = executor
            list(executor.map(process_record, all_records))
        return JsonResponse({})  
 
class YaserMarket(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        data = []
        for i in range(5000):
            newUrl = re.sub(r'page=\d+&','page='+str(i+1)+'&',url)
            response = requests.get(newUrl)
            products = response.json()['products']
            if len(products)==0:
                break
            for p in products:
                if p['stock_status']=='In Stock':
                    img = ''
                    image_url = 'https://www.icn.com/api/v1/image/upload'
                    data_to_upload = {
                        'user_id': request.data['user_id'],
                        'image_url': p['thumb']
                    }

                    try:
                        response = requests.post(image_url, data=data_to_upload)
                        if response.status_code == 200:
                            img= response.text
                    except requests.exceptions.RequestException as e:
                        print('An error occurred:', e)

                    data.append({
                        "Arabic Name": translate(p['name']),
                        "English Name": translate(p['name'], dest="en"),
                        "Arabic Description": translate(p['description']),
                        "English Description": translate(p['description'], dest="en"),
                        "Category Id": request.data['category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": p['price'],
                        "Discount Type": "",
                        "Discount": "",
                        "Unit": "PC",
                        "Current Stock": "3",
                        "Main Image URL": img,
                        "Photos URLs": img,
                        "Video Youtube URL": "",
                        "English Meta Tags": translate(p['name'], dest="en"),
                        "Arabic Meta Tags": translate(p['name']),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": 'https://www.yasermallonline.com/en/product/'+p['product_id'],
                    })
        df = pd.DataFrame([d for d in data if d['Current Stock'] != '0'])
        df.to_excel('excel/'+request.data['id']+'_products.xlsx', index=False)
        if os.path.join('excel', request.data['id']+'_products.xlsx'):
            try:
                url = "https://ai.icn.com/api/upload_image"
                files = {
                    'file': (request.data['id']+'_products.xlsx', open(os.path.join('excel', request.data['id']+'_products.xlsx'), 'rb'))  # Open the image in binary mode
                }
                data = {
                    'base_id': 'app4m95tHoPe9i69Y',
                    'table_id': 'tblC9J5CWTfclSbTD',
                    'record_id': request.data['id'],
                }
                # Send the POST request
                response = requests.post(url, files=files, data=data)
            except Exception as error:
                print(error)
        return JsonResponse({'data': []})

class Jubran(APIView):
    def post(self, request, *args, **kwargs):
        data = []
        for i in range(5000):
            print({
                "limit":24,
                "sort":"",
                "categoryId": request.data['categoryId'],
                "isOffer":False,
                "nextPageNumber":i,
                "pageSize":24,
                "isWholeSale":False
            })
            headers = {
                'Accept': 'application/json',
                'Accept-Language': 'en-US,en;q=0.9,ar;q=0.8',
                'Connection': 'keep-alive',
                'Content-Type': 'application/json',
                'Origin': 'https://jubranjo.com',
                'Referer': 'https://jubranjo.com/',
                'Sec-Fetch-Dest': 'empty',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Site': 'cross-site',
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
                'deviceType': 'desktopWeb',
                'sec-ch-ua': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"macOS"'
            }

            response = requests.post('https://jubran.jubran-api.com/api/products/searchProducts',data=json.dumps({
                "limit":24,
                "sort":"",
                "categoryId": request.data['categoryId'],
                "isOffer":False,
                "nextPageNumber":i,
                "pageSize":24,
                "isWholeSale":False
            }), headers=headers)
            print(response.text)
            products = response.json()['products']
            if len(products)==0:
                break
            for p in products:
                if p['isStockAvailable']== True:
                    img = ''
                    image_url = 'https://www.icn.com/api/v1/image/upload'
                    data_to_upload = {
                        'user_id': request.data['user_id'],
                        'image_url': p['imageUrl']
                    }

                    try:
                        response = requests.post(image_url, data=data_to_upload)
                        if response.status_code == 200:
                            img= response.text
                    except requests.exceptions.RequestException as e:
                        print('An error occurred:', e)

                    data.append({
                        "Arabic Name": 'جبران ' + translate(p['name']),
                        "English Name": 'Jubran ' + translate(p['name'], dest="en"),
                        "Arabic Description": request.data['ar_description'],
                        "English Description": request.data['en_description'],
                        "Category Id": request.data['category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": p['price'],
                        "Discount Type": "",
                        "Discount": "",
                        "Unit": "PC",
                        "Current Stock": "100",
                        "Main Image URL": img,
                        "Photos URLs": img,
                        "Video Youtube URL": "",
                        "English Meta Tags": translate(p['name'], dest="en"),
                        "Arabic Meta Tags": translate(p['name']),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": 'https://jubran.jubran-api.com/api/products/searchProducts/'+str(p['id']),
                    })
        df = pd.DataFrame([d for d in data if d['Current Stock'] != '0'])
        df.to_excel('excel/'+request.data['id']+'_products.xlsx', index=False)
        if os.path.join('excel', request.data['id']+'_products.xlsx'):
            try:
                url = "https://ai.icn.com/api/upload_image"
                files = {
                    'file': (request.data['id']+'_products.xlsx', open(os.path.join('excel', request.data['id']+'_products.xlsx'), 'rb'))  # Open the image in binary mode
                }
                data = {
                    'base_id': 'appiRnmkCqRxRVp1D',
                    'table_id': 'tblxoNzr4fst5S7BH',
                    'record_id': request.data['id'],
                }
                # Send the POST request
                response = requests.post(url, files=files, data=data)
            except Exception as error:
                print(error)
        return JsonResponse({'data': []})

class SecoundYaserMarket(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        data = []
        for i in range(1):
            newUrl = re.sub(r'page=\d+&','page='+str(i+1)+'&',url)
            response = requests.get(newUrl, headers={
                'Lang': 'en-gb',
                # 'Lang': 'ar'
            })
            products = response.json()['products']
            print(products[0])
            if len(products)==0:
                break
            for i,p in enumerate(products):
                if i > 2:
                    break
                if p['stock_status']=='In Stock':
                    img = p['thumb']
                    # img = ''
                    # image_url = 'https://www.icn.com/api/v1/image/upload'
                    # data_to_upload = {
                    #     'user_id': request.data['user_id'],
                    #     'image_url': p['thumb']
                    # }

                    # try:
                    #     response = requests.post(image_url, data=data_to_upload)
                    #     if response.status_code == 200:
                    #         img= response.text
                    # except requests.exceptions.RequestException as e:
                    #     print('An error occurred:', e)

                    data.append({
                        'ID': p['product_id'],
                        'Type': 'simple',
                        'SKU': '',
                        'Name': p['name'],
                        'Published': '-1',
                        'Is featured?': '0',
                        'Visibility in catalog': 'visible',
                        'Short description': '',
                        'Description': p['description'],
                        'Date sale price starts': '',
                        'Date sale price ends': '',
                        'Tax status': 'taxable',
                        'Tax class': '',
                        'In stock?': '1' if p['stock'] else '0',
                        'Stock': '',
                        'Low stock amount': '',
                        'Backorders allowed?': '0',
                        'Sold individually?': '0',
                        'Weight (kg)': '',
                        'Length (cm)': '',
                        'Width (cm)': '',
                        'Height (cm)': '',
                        'Allow customer reviews?': '1',
                        'Purchase note': '',
                        'Sale price': '',
                        'Regular price': p['price'],
                        'Categories': '',
                        'Tags': '',
                        'Shipping class': '',
                        'Images': p['thumb'],
                        'Download limit': '',
                        'Download expiry days': '',
                        'Parent': '',
                        'Grouped products': '',
                        'Upsells': '',
                        'Cross-sells': '',
                        'External URL': '',
                        'Button text': '',
                        'Position': '0',
                        'Language': 'ar',
                        'Translation group': '',
                        'Bundled Items (JSON-encoded)': '',
                        'Min Bundle Size': '',
                        'Max Bundle Size': '',
                        'Bundle Contents Virtual': '',
                        'Bundle Aggregate Weight': '',
                        'Bundle Layout': '',
                        'Bundle Group Mode': '',
                        'Bundle Cart Editing': '',
                        'Bundle Sold Individually': '',
                        'Bundle Form Location': '',
                        'Bundle Sells': '',
                        'Bundle Sells Title': '',
                        'Bundle Sells Discount': '',
                        'Meta: klb_product_badge_type': 'type1',
                        'Meta: _app_builder_shopping_video_addons_video_url': '',
                        'Meta: _app_builder_shopping_video_addons_video_name': '',
                        'Meta: _app_builder_shopping_video_addons_video_description': '',
                        'Meta: _klb_single_video_input': '',
                        'Meta: _klb_product_percentage_type': 'style-1',
                        'Meta: _klb_product_percentage_bg_color': '',
                        'Meta: _klb_product_percentage_text_color': '',
                        'Meta: _klb_product_badge_text': '',
                        'Meta: _klb_product_badge_type': 'style-1',
                        'Meta: _klb_product_badge_bg_color': '',
                        'Meta: _klb_product_badge_text_color': '',
                        'Meta: _secondary_title': '',

                        # "Arabic Name": translate(p['name']),
                        # "English Name": translate(p['name'], dest="en"),
                        # "reference_link": 'https://www.yasermallonline.com/en/product/'+p['product_id'],
                    })
        df = pd.DataFrame([d for d in data])
        df.to_excel('excel/'+request.data['id']+'_products.xlsx', index=False)
        # if os.path.join('excel', request.data['id']+'_products.xlsx'):
        #     try:
        #         url = "https://ai.icn.com/api/upload_image"
        #         files = {
        #             'file': (request.data['id']+'_products.xlsx', open(os.path.join('excel', request.data['id']+'_products.xlsx'), 'rb'))  # Open the image in binary mode
        #         }
        #         data = {
        #             'base_id': 'app4m95tHoPe9i69Y',
        #             'table_id': 'tblC9J5CWTfclSbTD',
        #             'record_id': request.data['id'],
        #         }
        #         # Send the POST request
        #         response = requests.post(url, files=files, data=data)
        #     except Exception as error:
        #         print(error)
        return JsonResponse({'data': []})
    
class StopProcess(APIView):
    def post(self, request, *args, **kwargs):
        global mainExecutor
        mainExecutor.shutdown()
        return JsonResponse({})

class InimexShopScrapView(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        isExist = True
        hrefs = []
        index=1
        fitst_index = index
        selector = ".product-panel > div > div.product-card-container"
        pagination = '?page='
        error = False
        while(isExist):
            if index != fitst_index:
                driver.get(url+pagination+str(index)+('/' if 'page=' not in pagination and 'pageNumber=' not in pagination and pagination != 'p=' else ''))
            sleep(3)
            isExist = check_if_exist(driver, selector, "products")
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            u = driver.current_url
            for i, e in enumerate(elements):
                driver.get(u)
                until_visible_click(driver, f'.product-panel > div > div.product-card-container:nth-child({i+1})')
                if not error:
                    try:
                        sleep(1)
                        title_selector = 'meta[property="og:title"]'
                        key_words_selector = "meta[property*='og:title']"
                        description_selector = 'meta[property="og:description"]'
                        try:
                            until_visible(driver, '.preview-container img.iiz__img')
                        except: 
                            pass
                        if len(driver.find_elements(By.CSS_SELECTOR, title_selector))==0:
                            continue
                        href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                        soup = BeautifulSoup(href_res, 'html.parser')
                        title = soup.select_one(title_selector)['content'].strip()
                        # Get the product price
                        price = soup.select_one('meta[property="og:price:amount"]')['content'].replace(',','').strip() if len(soup.select('meta[property="og:price:amount"]'))>0 else ''
                        # Get discount
                        discount = '0'
                        # Get the main image URL
                        main_image_elem = soup.select_one('.preview-container img.iiz__img')
                        image = getImageBase64(driver, request.data['id'], main_image_elem['src']) if main_image_elem else ''
                        # Get additional images
                        image_elems = soup.select('.preview-container img.iiz__img')
                        images = []
                        for img in image_elems:
                            if len(img['src'])>10:
                                res = getImageBase64(driver, request.data['id'], img['src'])
                                if res:
                                    images.append(res)
                        # Check stock status
                        in_stock = '3'
                        # Get product attributes content
                        description_elem = soup.select_one(description_selector)['content'].strip() if soup.select_one(description_selector) else ''
                        product_attributes_content = description_elem if description_elem else ''
                        # Get keywords
                        key_words_elem = soup.select_one(key_words_selector)
                        keyWords = key_words_elem['content'].strip() if key_words_elem else ''
                        keywords = keyWords.split('//')
                        if len(product_attributes_content)>0:
                            keywords = extract_top_keywords(product_attributes_content)
                            ar_keywords = []
                            for k in keyWords.split('//'):
                                keywords.append(k)

                            for keyW in keywords:
                                ar_keywords.append(translate(keyW))
                        else:
                            ar_keywords = []
                            for keyW in keywords:
                                ar_keywords.append(translate(keyW))
                        
                        product = {
                            "Arabic Name": translate(title),
                            "English Name": title,
                            "Arabic Description": translate(product_attributes_content) if len(product_attributes_content)>3 else request.data['arabic_description'],
                            "English Description": product_attributes_content if len(product_attributes_content) > 3 else request.data['description'],
                            "Category Id": request.data['db_category'],
                            "Arabic Brand": "",
                            "English Brand": "",
                            "Unit Price": price,
                            "Discount Type": "Flat" if discount != "0" else "",
                            "Discount": discount if discount != "0" else "",
                            "Unit": "PC",
                            "Current Stock": in_stock,
                            "Main Image URL": image,
                            "Photos URLs": str((",").join(images)) if images else image,
                            "Video Youtube URL": "",
                            "English Meta Tags": ','.join(keywords),
                            "Arabic Meta Tags": ','.join(ar_keywords),
                            "features": '',
                            "features_ar": '',
                            "wholesale": "no",
                            "reference_link": driver.current_url,
                        }
                        data.append(product)
                    except Exception as e:
                        error = True
                        print(e)
                        traceback.print_exc()
                        errors.append({
                            "url": driver.current_url
                        })   

            index = index + 1
                
            
        if len(errors)>0:
            err_df = pd.DataFrame(errors)
            err_df.to_excel('excel/'+request.data['db_category']+'_errors.xlsx', index=False)
        else:
            df = pd.DataFrame(data)
            df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
            change_content(driver, data, request.data['db_category'])
            
        driver.quit()
        return JsonResponse({})

class GetProductsFromBing(APIView):
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        store_id = request.data['id']

        if not file or not isinstance(file, InMemoryUploadedFile):
            return JsonResponse({"error": "No file provided or invalid file"})

        # Read the file and load it into openpyxl
        file_content = file.read()
        file_name = file.name
        print(file_name)
        wb = load_workbook(filename=io.BytesIO(file_content))
        sheet = wb.active

        # Process the Excel file (example: read cell values)
        headers = [str(cell.value).strip() for cell in sheet[1]]

        # Process the Excel file and format data as an array of objects
        excel_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue  # Skip empty rows
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            row_data['photo'] = ''
            row_data['description'] = ''
            excel_data.append(row_data)

        def getImage(d):
            driver = create_browser()
            try:
                # Open Google Images in the browser
                driver.get('https://www.bing.com/search?q='+d['item_name'])

                # Finding the search box
                hrefs = [d.get_attribute('href') for d in driver.find_elements(By.CSS_SELECTOR,
                                             'ol > li > div > a[href]')]
                for indx, h in enumerate(hrefs):
                    driver.get(h)
                    if len(d['photo'])>0 and len(d['description'])>0:
                        break
                    if len(driver.find_elements(By.CSS_SELECTOR, 'meta[property="og:image"]'))>0 and len(d['photo'])==0:
                        imgSrc = driver.find_element(By.CSS_SELECTOR, 'meta[property="og:image"]').get_attribute('content')
                        d['photo'] = getImageBase64(driver, store_id, imgSrc)

                    if len(driver.find_elements(By.CSS_SELECTOR, 'meta[property="twitter:image"]'))>0 and len(d['photo'])==0:
                        imgSrc = driver.find_element(By.CSS_SELECTOR, 'meta[property="twitter:image"]').get_attribute('content')
                        d['photo'] = getImageBase64(driver, store_id, imgSrc)

                    if len(driver.find_elements(By.CSS_SELECTOR, 'meta[property="og:description"]'))>0 and len(d['description'])==0:
                        description = driver.find_element(By.CSS_SELECTOR, 'meta[property="og:description"]').get_attribute('content')
                        d['description'] = description

                    if len(driver.find_elements(By.CSS_SELECTOR, 'meta[property="description"]'))>0 and len(d['description'])==0:
                        description = driver.find_element(By.CSS_SELECTOR, 'meta[property="description"]').get_attribute('content')
                        d['description'] = description

            except:
                pass
            finally:
                driver.close()
        with ThreadPoolExecutor(max_workers=6) as executor:
            list(executor.map(getImage, [d for d in excel_data]))
        data = []
        for e in excel_data:
            product = {
                "Arabic Name": e['item_localized_name'],
                "English Name": e['item_name'],
                "Arabic Description": translate(e['description']) if len(e['description']) > 0 else '',
                "English Description": e['description'],
                "Category Id": "",
                "Arabic Brand": "",
                "English Brand": "",
                "Unit Price": e['item_price'],
                "Discount Type": "",
                "Discount": "",
                "Unit": "PC",
                "Current Stock": 3,
                "Main Image URL": e['photo'],
                "Photos URLs": str((",").join([e['photo']])) if e['photo'] else "",
                "Video Youtube URL": "",
                "English Meta Tags": "",
                "Arabic Meta Tags": "",
                "features": '',
                "wholesale": "no",
                "reference_link": "",
            }
            data.append(product)
        df = pd.DataFrame(data)
        df.to_excel('excel/'+file_name.replace('.xlsx','')+'.xlsx', index=False)

        return JsonResponse({})

class GetImagesFromGoogle(APIView):
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        store_id = request.data['id']

        if not file or not isinstance(file, InMemoryUploadedFile):
            return JsonResponse({"error": "No file provided or invalid file"})
        
        # Read the file and load it into openpyxl
        file_content = file.read()
        file_name = file.name
        print(file_name)
        wb = load_workbook(filename=io.BytesIO(file_content))
        sheet = wb.active
        
        # Process the Excel file (example: read cell values)
        headers = [str(cell.value).strip() for cell in sheet[1]]
        
        # Process the Excel file and format data as an array of objects
        excel_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue  # Skip empty rows
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            excel_data.append(row_data)
        
        def getImage(d):
            driver = create_browser()
            try:
                # Open Google Images in the browser
                driver.get('https://images.google.com/')
                
                # Finding the search box
                box = driver.find_element(By.XPATH,'//*[@action="https://www.google.com/search"]/div/div/div/div/div[2]/textarea')        
                box.send_keys(d['وصف المادة'])        
                box.send_keys(Keys.ENTER)
                until_visible_with_xpath(driver, '//*[@id="search"]/div[1]/div/div/div/div[1]/div/div/div//h3/a')
                hrefs = driver.find_elements(By.XPATH,
                    '//*[@id="search"]/div[1]/div/div/div/div[1]/div/div/div//h3/a')
                url = driver.current_url
                for indx, h in enumerate(hrefs):
                    driver.get(url)
                    until_visible_xpath_click(driver, '//*[@id="search"]/div[1]/div/div/div/div[1]/div/div/div['+str(indx + 1)+']//h3/a')
                    sleep(0.5)
                    until_visible_with_xpath(driver, "//span[text()='انتقال']")
                    driver.get(driver.find_element(By.XPATH, "//span[text()='انتقال']/parent::*/parent::*").get_attribute('href'))
                    sleep(5)
                    if len(driver.find_elements(By.CSS_SELECTOR, 'meta[property="og:image"]'))>0:
                        imgSrc = driver.find_element(By.CSS_SELECTOR, 'meta[property="og:image"]').get_attribute('content')
                        d['photo'] = getImageBase64(driver, store_id, imgSrc)
                        break
            except:
                pass
            finally:
                driver.close()
        with ThreadPoolExecutor(max_workers=6) as executor:
            list(executor.map(getImage, [d for d in excel_data if d['photo'] == None]))
        driver = create_browser()
        driver.get('https://images.google.com/')
        data = []
        for e in excel_data:
            try:
                if e['photo'] != None and e['photo'] != '' and 'icn.com' not in e['photo']:
                    driver.get(e['photo'])
                    image_link = getImageBase64(driver, store_id, e['photo'])
                    e['photo'] = image_link
            except:
                e['photo'] = ''
            product = {
                    "Arabic Name": translate(e['وصف المادة']),
                    "English Name": e['وصف المادة'],
                    "Arabic Description": translate(e['وصف المادة']),
                    "English Description": e['وصف المادة'],
                    "Category Id": "",
                    "Arabic Brand": "",
                    "English Brand": "",
                    "Unit Price": e['price'],
                    "Discount Type": "",
                    "Discount": "",
                    "Unit": "PC",
                    "Current Stock": 3,
                    "Main Image URL": e['photo'],
                    "Photos URLs": str((",").join([e['photo']])) if e['photo'] else "",
                    "Video Youtube URL": "",
                    "English Meta Tags": "",
                    "Arabic Meta Tags": "",
                    "features": '',
                    "wholesale": "no",
                    "reference_link": "",
                }
            data.append(product)
        df = pd.DataFrame(data)
        df.to_excel('excel/'+file_name+'.xlsx', index=False)
        
        return JsonResponse({})

class CommonWebsites(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        print(url)
        website = Websites.objects.all().order_by('-id')
        driver = create_browser()
        driver.get(url)
        sleep(3)
        for web in website:
            if web.inner_selector and len(driver.find_elements(By.CSS_SELECTOR, web.product_selector))>0:
                print('1')
                if len(driver.find_element(By.CSS_SELECTOR, web.product_selector).find_elements(By.CSS_SELECTOR, web.inner_selector))>0:
                    print('2')
                    driver.get(driver.find_element(By.CSS_SELECTOR, web.product_selector).find_element(By.CSS_SELECTOR, web.inner_selector).get_attribute('href'))
                    sleep(3)
                    if len(driver.find_elements(By.CSS_SELECTOR, web.title_selector))>0 and len(driver.find_elements(By.CSS_SELECTOR, web.main_img_selector))>0 and len(driver.find_elements(By.CSS_SELECTOR, web.img_selector))>0 and len(driver.find_elements(By.CSS_SELECTOR, web.price_selector))>0:
                        print('3')
                        print(web.product_selector)
                        newWebsite = Websites()
                        newWebsite.require_login = web.require_login
                        newWebsite.email_selector = web.email_selector
                        newWebsite.email = web.email
                        newWebsite.password_selector = web.password_selector
                        newWebsite.password = web.password
                        newWebsite.button_selector = web.button_selector
                        newWebsite.no_pagination = web.no_pagination
                        newWebsite.pagination_click = web.pagination_click
                        newWebsite.pagination_path = web.pagination_path
                        newWebsite.product_selector = web.product_selector
                        newWebsite.not_contains_class = web.not_contains_class
                        newWebsite.inner_selector = web.inner_selector
                        newWebsite.inside_category_selector = web.inside_category_selector
                        newWebsite.product_click = web.product_click
                        newWebsite.title_prefix = web.title_prefix
                        newWebsite.title_prefix_selector = web.title_prefix_selector
                        newWebsite.title_prefix_attr = web.title_prefix_attr
                        newWebsite.title_selector = web.title_selector
                        newWebsite.title_attr = web.title_attr
                        newWebsite.title_suffix = web.title_suffix
                        newWebsite.title_suffix_selector = web.title_suffix_selector
                        newWebsite.title_suffix_attr = web.title_suffix_attr
                        newWebsite.description_selector = web.description_selector
                        newWebsite.description_attr = web.description_attr
                        newWebsite.key_words_selector = web.key_words_selector
                        newWebsite.main_img_selector = web.main_img_selector
                        newWebsite.main_img_attr = web.main_img_attr
                        newWebsite.img_click = web.img_click
                        newWebsite.img_selector = web.img_selector
                        newWebsite.img_attr = web.img_attr
                        newWebsite.static_price = web.static_price
                        newWebsite.is_price_have_comma = web.is_price_have_comma
                        newWebsite.price_selector = web.price_selector
                        newWebsite.price_attr = web.price_attr
                        newWebsite.second_price_selector = web.second_price_selector
                        newWebsite.second_price_attr = web.second_price_attr
                        newWebsite.is_discount = web.is_discount
                        newWebsite.discount_selector = web.discount_selector
                        newWebsite.discount_attr = web.discount_attr
                        newWebsite.is_stuck = web.is_stuck
                        newWebsite.stuck_selector = web.stuck_selector
                        newWebsite.is_feature = web.is_feature
                        newWebsite.features_selector = web.features_selector
                        newWebsite.features_key_selector = web.features_key_selector
                        newWebsite.features_key_attr = web.features_key_attr
                        newWebsite.features_value_selector = web.features_value_selector
                        newWebsite.features_value_attr = web.features_value_attr
                        newWebsite.en_link = web.en_link
                        newWebsite.ar_link = web.ar_link
                        newWebsite.ar_selector = web.ar_selector
                        newWebsite.ar_attr = web.ar_attr
                        newWebsite.export_out_of_stuck = web.export_out_of_stuck
                        newWebsite.start_index = web.start_index
                        newWebsite.end_index = web.end_index
                        newWebsite.number_of_products = web.number_of_products
                        newWebsite.change_content = web.change_content
                        newWebsite.save()
                        return JsonResponse({
                            'product_selector': web.product_selector,
                            'title_selector': web.title_selector,
                            'img_selector': web.img_selector,
                            'price_selector': web.price_selector
                        })
                    else:
                        driver.get(url)
                        sleep(3)
                else:
                    driver.get(url)
                    sleep(3)
            elif len(driver.find_elements(By.CSS_SELECTOR, web.product_selector))>0:
                print( web.name)
                print( web.product_selector)
                print( web.inner_selector)
                print( web.inside_category_selector)
                try:
                    driver.get(driver.find_element(By.CSS_SELECTOR, web.product_selector).get_attribute('href'))
                    sleep(3)
                    if len(driver.find_elements(By.CSS_SELECTOR, web.title_selector))>0 and len(driver.find_elements(By.CSS_SELECTOR, web.main_img_selector))>0 and len(driver.find_elements(By.CSS_SELECTOR, web.img_selector))>0 and len(driver.find_elements(By.CSS_SELECTOR, web.price_selector))>0:
                        print('2')
                        print(web.product_selector)
                        newWebsite = Websites()
                        newWebsite.require_login = web.require_login
                        newWebsite.email_selector = web.email_selector
                        newWebsite.email = web.email
                        newWebsite.password_selector = web.password_selector
                        newWebsite.password = web.password
                        newWebsite.button_selector = web.button_selector
                        newWebsite.no_pagination = web.no_pagination
                        newWebsite.pagination_click = web.pagination_click
                        newWebsite.pagination_path = web.pagination_path
                        newWebsite.product_selector = web.product_selector
                        newWebsite.not_contains_class = web.not_contains_class
                        newWebsite.inner_selector = web.inner_selector
                        newWebsite.inside_category_selector = web.inside_category_selector
                        newWebsite.product_click = web.product_click
                        newWebsite.title_prefix = web.title_prefix
                        newWebsite.title_prefix_selector = web.title_prefix_selector
                        newWebsite.title_prefix_attr = web.title_prefix_attr
                        newWebsite.title_selector = web.title_selector
                        newWebsite.title_attr = web.title_attr
                        newWebsite.title_suffix = web.title_suffix
                        newWebsite.title_suffix_selector = web.title_suffix_selector
                        newWebsite.title_suffix_attr = web.title_suffix_attr
                        newWebsite.description_selector = web.description_selector
                        newWebsite.description_attr = web.description_attr
                        newWebsite.key_words_selector = web.key_words_selector
                        newWebsite.main_img_selector = web.main_img_selector
                        newWebsite.main_img_attr = web.main_img_attr
                        newWebsite.img_click = web.img_click
                        newWebsite.img_selector = web.img_selector
                        newWebsite.img_attr = web.img_attr
                        newWebsite.static_price = web.static_price
                        newWebsite.is_price_have_comma = web.is_price_have_comma
                        newWebsite.price_selector = web.price_selector
                        newWebsite.price_attr = web.price_attr
                        newWebsite.second_price_selector = web.second_price_selector
                        newWebsite.second_price_attr = web.second_price_attr
                        newWebsite.is_discount = web.is_discount
                        newWebsite.discount_selector = web.discount_selector
                        newWebsite.discount_attr = web.discount_attr
                        newWebsite.is_stuck = web.is_stuck
                        newWebsite.stuck_selector = web.stuck_selector
                        newWebsite.is_feature = web.is_feature
                        newWebsite.features_selector = web.features_selector
                        newWebsite.features_key_selector = web.features_key_selector
                        newWebsite.features_key_attr = web.features_key_attr
                        newWebsite.features_value_selector = web.features_value_selector
                        newWebsite.features_value_attr = web.features_value_attr
                        newWebsite.en_link = web.en_link
                        newWebsite.ar_link = web.ar_link
                        newWebsite.ar_selector = web.ar_selector
                        newWebsite.ar_attr = web.ar_attr
                        newWebsite.export_out_of_stuck = web.export_out_of_stuck
                        newWebsite.start_index = web.start_index
                        newWebsite.end_index = web.end_index
                        newWebsite.number_of_products = web.number_of_products
                        newWebsite.change_content = web.change_content
                        newWebsite.save()
                        return JsonResponse({
                            'product_selector': web.product_selector,
                            'title_selector': web.title_selector,
                            'img_selector': web.img_selector,
                            'price_selector': web.price_selector
                        })
                    else:
                        driver.get(url)
                        sleep(3)
                except:
                    pass
        # sleep(10)
        
        driver.quit()
        return JsonResponse({})

class HyperMax(APIView):
    def post(self, request, *args, **kwargs):
        url = 'https://api-prod.retailsso.com/v3/categories/mafjor/ar/categories/'+request.data['url']+'?deviceId=B8E1E7BF-068F-4026-8B30-4228D40CCF0D&disableSpellCheck=true&nextOffset=0&needVariantsData=false&maxPrice=&sortBy=relevance&minPrice=&lang=ar&filter=&pageSize=40&currentPage=0&longitude=35.8354861&latitude=31.9804777'
        data = []
        driver = create_browser()
        driver.get('https://api-prod.retailsso.com/')
        for i in range(5000):
            newUrl = re.sub(r'currentPage=\d+&','currentPage='+str(i)+'&',url)
            response = requests.get(newUrl, headers={
                'userid': 'B8E1E7BF-068F-4026-8B30-4228D40CCF0D',
                'appid': 'IOS',
                'storeid': 'mafjor',
                'user-agent': 'MAFCarrefour/Prod-24.11.15(1)(iOS 16.7.10)',
                'nbinfo': 'false',
                'currency': 'JOD',
                'appversion': '241115',
                'appflavour': 'hypermax',
                'newrelic': 'ewoiZCI6IHsKImFjIjogIjMzNTU3MjAiLAoiYXAiOiAiMTEzMzkwODg3MyIsCiJpZCI6ICJjMmI3NGIzMTU1YWUyNTYwIiwKInRpIjogMTczNjE1MzE1NjQ2MiwKInRyIjogIjdjMDFhYzI5YWY0MGFlMTA5ZjVjN2ZhZDgyNTc2OWI4IiwKInR5IjogIk1vYmlsZSIKfSwKInYiOiBbCjAsCjIKXQp9',
                'deviceid': 'B8E1E7BF-068F-4026-8B30-4228D40CCF0D',
                'producttype': 'ANY',
                'token': '',
                'env': 'PROD',
                'tracestate': '@nr=0-2-3355720-1133908873-c2b74b3155ae2560--0--1736153156462',
                'accept-language': 'ar',
                'osversion': '16.7.10',
                'posinfo': 'food=040_Zone01,nonfood=040_Zone01,express=040_Zone01',
                'intent': 'STANDARD',
                'content-type': 'application/json',
                'traceparent': '00-7c01ac29af40ae109f5c7fad825769b8-c2b74b3155ae2560-00',
                'x-newrelic-id': 'VwUCVFFRCBABVVJRDgEPXlMC',
                'langcode': 'ar'
            })
            print(newUrl)
            print(response.json())
            products = response.json()['data']['products']
            if len(products)==0:
                break
            for p in products:
                if p['stock']['stockLevelStatus']=='inStock':
                    img = ''
                    img = getImageBase64(driver, request.data['user_id'],  p['links']['images'][0]['href'])
                    # image_url = 'https://www.icn.com/api/v1/image/upload'
                    # data_to_upload = {
                    #     'user_id': request.data['user_id'],
                    #     'image_url': p['links']['images'][0]['href']
                    # }
                    print(p['links']['images'][0]['href'])
                    print(img)

                    # try:
                    #     response = requests.post(image_url, data=data_to_upload)
                    #     print(response)
                    #     if response.status_code == 200:
                    #         img= response.text
                    # except requests.exceptions.RequestException as e:
                    #     print('An error occurred:', e)

                    data.append({
                        "Arabic Name": translate(p['name']),
                        "English Name": translate(p['name'], dest="en"),
                        "Arabic Description": p['name'],
                        "English Description": p['name'],
                        "Category Id": request.data['category'],
                        "Arabic Brand": "",
                        "English Brand": "",
                        "Unit Price": p['price']['price'],
                        "Discount Type": "",
                        "Discount": "",
                        "Unit": "PC",
                        "Current Stock": "99",
                        "Main Image URL": img,
                        "Photos URLs": img,
                        "Video Youtube URL": "",
                        "English Meta Tags": translate(p['name'], dest="en"),
                        "Arabic Meta Tags": translate(p['name']),
                        "features": '',
                        "features_ar": '',
                        "wholesale": "no",
                        "reference_link": p['links']['productUrl']['href'],
                    })
        df = pd.DataFrame([d for d in data if d['Current Stock'] != '0'])
        df.to_excel('excel/'+request.data['id']+'_products.xlsx', index=False)
        if os.path.join('excel', request.data['id']+'_products.xlsx'):
            try:
                url = "https://ai.icn.com/api/upload_image"
                files = {
                    'file': (request.data['id']+'_products.xlsx', open(os.path.join('excel', request.data['id']+'_products.xlsx'), 'rb'))  # Open the image in binary mode
                }
                data = {
                    'base_id': 'appiRnmkCqRxRVp1D',
                    'table_id': 'tblhSLXdzuMlAp0lq',
                    'record_id': request.data['id'],
                }
                # Send the POST request
                response = requests.post(url, files=files, data=data)
            except Exception as error:
                print(error)
        return JsonResponse({'data': []})

def getTexts(driver, tex):
  print('1')
  res = []
  try:
    # data = {"synset":tex,"level":2,"useALMA":True}
    # data = json.dumps(data).encode('utf8')
          
    # req = Request(
    #   url="https://ontology.birzeit.edu/sina/v2/api/SynonymGenerator/?apikey=sinaMapping",
    #   data=data,
    #   headers={'User-Agent': 'Mozilla/5.0', 'Content-Type': 'application/json'},
    #   method='POST'
    # )
    # webpage = urlopen(req).read()
    webpage = sendRequest(driver, "https://ontology.birzeit.edu/sina/v2/api/SynonymGenerator/?apikey=sinaMapping", {"synset":tex,"level":2,"useALMA":True})
    print(webpage['resp'])
    diacritics_pattern = re.compile(r"[\u064B-\u065F]")
    for d in webpage['resp']:
      cleaned_word = re.sub(diacritics_pattern, '', d[0])
      res.append(cleaned_word)
  except Exception as e:
    print(e)
    pass
  res.append(tex)
  return res

def is_arabic_word(value):
    if isinstance(value, str):
        # Check if it contains Arabic characters and is not purely numeric
        return bool(re.search(r'[\u0600-\u06FF]', value)) and not value.isdigit()
    return False
    
class SimilarNames(APIView):
    def post(self, request, *args, **kwargs):
        driver = create_browser()
        driver.get('https://sina.birzeit.edu/synonyms/')
        texts = []
        i=1
        while(True):
            print('getting data')
            res = requests.get('https://www.icn.com/api/v1/products/get?id='+request.data['id']+'&page='+str(i))
            products = res.json()['data']['products']
            for p in products:
                texts.append({'id': p['id'], 'name': p['name']})
            if len(res.json()['data']['products'])==0:
                break
            i=i+1
        print('finished')
        for t in texts:
            try:
                existProduct = Products.objects.filter(name=t['name']).get()
            except:
                existProduct = None
            if existProduct:
                continue
            print('getting words')
            result = []
            result.append(t['name'])
            for i, char in enumerate(t['name'].split(' ')):
                try:
                    existWord = Words.objects.filter(name=char).get()
                except:
                    existWord = None
                if existWord:
                    spilated = existWord.similarWords.split(',')
                    for s in spilated:
                        test= t['name'].split(' ')
                        test[i] = s
                        result.append(' '.join(test))
                else:
                    print(char)
                    print(is_arabic_word(char))
                    if is_arabic_word(char):
                        spilated = getTexts(driver, char)
                        for s in spilated:
                            test= t['name'].split(' ')
                            test[i] = s
                            result.append(' '.join(test))
                        newWords = Words()
                        newWords.name = char
                        newWords.similarWords = ','.join(spilated)
                        newWords.save()
            newProduct = Products()
            newProduct.id = t['id']
            newProduct.name = t['name']
            newProduct.similarNames = ','.join(result)
            newProduct.save()
        print('finished')
        driver.quit()
        return JsonResponse({})

# change category id
# electronic: 1
# fashion: 176
# pets: 485
# i=1
# while(True):
#     print('getting data')
#     res = requests.get('https://www.icn.com/api/v1/products/get?id='+'487'+'&page='+str(i))
#     products = res.json()['data']['products']
#     print(str(i)+'/'+str(res.json()['data']['links']['total_pages']))
#     for p in products:
#         try:
#             prod = Products.objects.get(id=p['id'])
#             prod.categoryId = 1588
#             prod.save()
#             print(prod.name)
#         except:
#             pass
#     if len(res.json()['data']['products'])==0:
#         break
#     i=i+1

import numpy as np
from numpy import dot
from numpy.linalg import norm

def preprocess_text(text):
    text = re.sub(r'[إأآا]', 'ا', text)  # Normalize Alef
    text = re.sub(r'ة', 'ه', text)  # Normalize Ta Marbuta
    text = re.sub(r'[يى]', 'ي', text)  # Normalize Ya
    text = re.sub(r'\d+', '', text)  # Remove digits
    text = re.sub(r'\b(ظرف|جرام|كيلو|مل)\b', '', text)  # Remove units
    text = re.sub(r'[^\w\s]', '', text)  # Remove punctuation
    return text.strip()

# productNames = []
# for p in Products.objects.filter(categoryId__isnull=False).order_by('-categoryId')[:50]:
    newNames = []
    newNames.append(p.name)
    for name in p.similarNames.split(','):
        if name != p.name:
            newNames.append(name)
    p.similarNames = ','.join(newNames)
    splittedNames = []
    for name in newNames:
        for i, n in enumerate(name.split(' ')):
            collectedName = []
            if n not in splittedNames:
                if i > 0:
                    # splittedNames.append(' '.join(name.split(' ')[:i])+' '+n)
                    for j in range(len(n)):
                        collectedName.append(n[j])
                        print(' '.join(name.split(' ')[:i])+' '+''.join(collectedName))
                        splittedNames.append(' '.join(name.split(' ')[:i])+' '+''.join(collectedName))
                else:
                    for j in range(len(n)):
                        collectedName.append(n[j])
                        print(collectedName)
                        splittedNames.append(''.join(collectedName))
                    # splittedNames.append(n)
    p.splittedNames = ','.join(splittedNames)
    print(p.id)
    p.save()

productNames = []
for p in Words.objects.all():
    p.delete()
# Train
# arabert_model = models.Transformer('aubmindlab/bert-base-arabert')
# pooling_layer = models.Pooling(arabert_model.get_word_embedding_dimension(), pooling_mode_mean_tokens=True)

# # Combine the transformer and pooling into a SentenceTransformer model
# model = SentenceTransformer(modules=[arabert_model, pooling_layer])
# # Prepare product names
# productNames = []
# for p in Products.objects.filter(categoryId__isnull=False).order_by('-categoryId')[:50]:
#     for name in p.splittedNames.split(','):
#         productNames.append({'id': p.id, 'name': name, 'originalName': p.name})
#     print(p.name)

# # Encode corpus
# corpus = [preprocess_text(r['name']) for r in productNames]
# print(len(corpus))
# # corpus_embeddings = model.encode(corpus, convert_to_tensor=True).detach().cpu().numpy()

# corpus_embeddings = []
# for i in range(0, len(corpus), 2048):
#     batch = corpus[i:i+2048]
#     print(str(i+2048)+'/'+str(len(corpus)))
#     embeddings = model.encode(batch, convert_to_tensor=True).detach().cpu().numpy()
#     corpus_embeddings.append(embeddings)
# corpus_embeddings = np.vstack(corpus_embeddings)

# # Create FAISS index
# dimension = corpus_embeddings.shape[1]
# wordIndex = faiss.IndexFlatL2(dimension)
# wordIndex.train(corpus_embeddings) 
# wordIndex.add(corpus_embeddings)

# ## save
# # Save FAISS index
# faiss.write_index(wordIndex, 'faiss_index.index')

# # Save corpus embeddings to a file
# np.save('corpus_embeddings.npy', corpus_embeddings)

# # Save the SentenceTransformer model
# model.save('sentence_transformer_model')

# # ## load
# # Load the SentenceTransformer model
# model = SentenceTransformer('sentence_transformer_model')

# # Load the FAISS index
# wordIndex = faiss.read_index('faiss_index.index')

# # Load the corpus embeddings
# corpus_embeddings = np.load('corpus_embeddings.npy')

# # train new
# new_embeddings = []
# for i in range(0, len(corpus), 2048):
#     batch = corpus[i:i+2048]
#     print(str(i+2048)+'/'+str(len(corpus)))
#     embeddings = model.encode(batch, convert_to_tensor=True).detach().cpu().numpy()
#     new_embeddings.append(embeddings)

# new_embeddings = np.vstack(new_embeddings)

# # Step 3: Update the FAISS index with the new embeddings
# # Ensure that the dimensionality matches
# assert new_embeddings.shape[1] == corpus_embeddings.shape[1], "Dimensionality of embeddings must match."

# # Add the new embeddings to the existing FAISS index
# wordIndex.add(new_embeddings)

# # Step 4: Optionally save the updated FAISS index
# faiss.write_index(wordIndex, 'faiss_index.index')

# # Step 5: Save the new corpus embeddings if needed
# np.save('corpus_embeddings.npy', new_embeddings)

# # If you need to append the new embeddings to the old ones
# updated_corpus_embeddings = np.vstack([corpus_embeddings, new_embeddings])
# np.save('corpus_embeddings.npy', updated_corpus_embeddings)
# model.save('sentence_transformer_model')

# Compute cosine similarity
def compute_cosine_similarity(embedding1, embedding2):
    return dot(embedding1, embedding2) / (norm(embedding1) * norm(embedding2))

# Token overlap score
def compute_jaccard_similarity(query, text):
    query_set = set(query.split())
    text_set = set(text.split())
    intersection = len(query_set & text_set)
    union = len(query_set | text_set)
    return intersection / union if union != 0 else 0
def normalize_distance(distance):
    return 1 / (1 + distance)

class Search(APIView):
    def post(self, request, *args, **kwargs):
        res = []
        query = request.data['query']
        query_processed = preprocess_text(query)
        query_embedding = model.encode(query_processed)
        query_embedding = np.array(query_embedding).astype('float32').reshape(1, -1)
        k = 10
        distances, indices = wordIndex.search(query_embedding, k)
        print(f"Query: {query_processed}")
        for i in range(k):
            # print(f"i: {i}, indices: {indices[0][0]}, productNames: {productNames}")
            product = productNames[indices[0][i]]
            product_name = product['name']
            product_name_processed = preprocess_text(product_name)

            # Calculate cosine similarity
            # Calculate Jaccard similarity for token overlap
            product_embedding = corpus_embeddings[indices[0][i]]
            cosine_similarity = compute_cosine_similarity(query_embedding[0], product_embedding)

            jaccard_similarity = compute_jaccard_similarity(query_processed, product_name_processed)

            # Normalize FAISS distance to similarity
            distance_similarity = normalize_distance(distances[0][i])

            # Final combined score
            final_score = (
                0.6 * distance_similarity +
                0.2 * cosine_similarity +
                0.2 * jaccard_similarity
            )
            # Add to results if not already included
            if product['id'] not in [p['id'] for p in res]:
                res.append({
                    'id': product['id'],
                    'originalName': product['originalName'],
                    'name': product_name,
                    'score': float(final_score),  # Convert to Python float
                    'cosine_similarity': float(cosine_similarity),  # Convert to Python float
                    'jaccard_similarity': float(jaccard_similarity),  # Convert to Python float
                    'distance': float(distances[0][i])  # Convert to Python float
                })
        MIN_SCORE_THRESHOLD = 0.2

        res = sorted(res, key=lambda x: x['score'], reverse=True)
        filtered_results = [result for result in res if result['score'] >= MIN_SCORE_THRESHOLD]

        print(res)
        return JsonResponse({'data': filtered_results})
     
def get_bsf4_xpath(element, multi=False):
    path_parts = []
    
    # Traverse upwards from the element to the root
    while element:
        siblings = element.find_previous_siblings(element.name)
        index = len(siblings) + 1
        path_parts.insert(0, f"{element.name}[{index}]")
        element = element.find_parent()
        if element.name == 'html':
            break

    if multi:
        last_bracket_index = '/'.join(path_parts).rfind('[')
        fixed_path = '/'.join(path_parts)[:last_bracket_index]
    else:
        fixed_path = '/'.join(path_parts)

    return '//' + fixed_path

def get_class_css_selector(element):
    path_parts = []

    while element:
        class_attr = element.get('class')
        
        if class_attr:
            # If the element has a class, build the CSS selector
            class_value = " ".join(class_attr)  # Combine multiple class names if necessary
            # Create the CSS selector for the class
            path_parts.insert(0, f".{class_value}")
        else:
            # If no class, use the tag name
            path_parts.insert(0, element.name)

        # Move to the parent element
        element = element.find_parent()

        if element and element.name == 'html':  # Stop at the root element
            break

    # Join all the path parts and use a space (descendant combinator) instead of '>'
    return ' '.join(path_parts)

def getBiggestTitle(driver):
    res = ''
    prices = []
    if len(driver.find_elements(By.CSS_SELECTOR, 'h1, h2, h3, h4, h5, h6, a, p, span, name'))>0:
        for price in driver.find_elements(By.CSS_SELECTOR, 'h1, h2, h3, h4, h5, h6, a, p, span, name'):
            if len(re.findall(r"\d{1,3}\.\d{1,3}", price.text))>0:
                continue
            if len(price.text)<5:
                continue
            font_weight = price.value_of_css_property('font-weight')  
            print(price.text)      
            print(font_weight)      
            if font_weight == 'normal':
                font_weight_value = 400
            elif font_weight == 'bold':
                font_weight_value = 700
            else:
                font_weight_value = int(font_weight)
            prices.append({
                'element': price,
                'size': font_weight_value,
            })
    biggest_size = 0
    for p in prices:
        if p['size'] > biggest_size:
            biggest_size = p['size']
            res = p['element'].text
    return res

def saveElementScreenShot(driver, element_selector, name):
    if element_selector:
        elements = driver.find_elements(By.XPATH, element_selector)
        for element in elements:
            driver.execute_script("""
                arguments[0].style.border = '5px solid red !important';
            """, element)
        screenshot_path = name+".png"
        driver.save_screenshot('imgs/'+screenshot_path)
        for element in elements:
            driver.execute_script("""
                arguments[0].style.border = '';
            """, element)
    

class Test(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        last_bracket_index = url.rfind('.com/')
        domain = url[:last_bracket_index]+'.com/'
        driver = create_browser()
        driver.get(url)
        sleep(10)
        data = []
        def checkIfExist(calsses, newClass):
            res = False
            for c in calsses:
                for n in newClass:
                    if str(c).strip() == str(n).strip():
                        res = True
            return res 
        index = 0
        
        # bs4
        # href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
        # soup = BeautifulSoup(href_res, 'html.parser')
        # for divs in soup.select('div,ul'):
        #     child_divs = divs.find_all(recursive=False)
        #     if len(child_divs)>3:
        #         try:
        #             print(child_divs[0].select('img'))
        #             first_div_height = child_divs[0]['class']
        #             if len(child_divs[0].select('img'))==0:
        #                 continue
        #             if first_div_height:
        #                 all_same_size = all(checkIfExist(first_div_height, div['class']) for div in child_divs)
        #                 if all_same_size:
        #                     hrefs = []
                            
        #                     for w in child_divs:
        #                         if len(w.select('a[href]'))>0:
        #                             href = w.select_one('a[href]')['href']
        #                             if 'http' not in href:
        #                                 hrefs.append(domain+href)
        #                             else:
        #                                 hrefs.append(href)
        #                     print('1')
        #                     xpath = get_bsf4_xpath(child_divs[0], True)
        #                     # css = get_class_css_selector(child_divs[0])
        #                     print(xpath)
        #                     data.append({
        #                         # 'class': ' '.join(divs['class']),
        #                         'name': 'Div: '+xpath,
        #                         'hrefs': hrefs
        #                     })
        #                     index = index + 1
        #         except Exception as e:
        #             print(e)
        
        # selenium
        for divs in driver.find_elements(By.CSS_SELECTOR,'div, ul, section'):
            selenium_child_divs = divs.find_elements(By.XPATH, './div | ./li')
            all_same_size = False
            if len(selenium_child_divs)>3:
                # for class
                first_div_height = selenium_child_divs[0].get_attribute('class')
                if first_div_height:
                    all_same_size = all(checkIfExist(first_div_height, div.get_attribute('class')) for div in selenium_child_divs)
                    if all_same_size:
                        hrefs = []
                        for w in selenium_child_divs:
                            if len(w.find_elements(By.CSS_SELECTOR,'a[href]'))>0:
                                href = w.find_element(By.CSS_SELECTOR,'a[href]').get_attribute('href')
                                title = getBiggestTitle(w)
                                if len(title.strip())>0:
                                    hrefs.append({
                                        'name': title,
                                        'link': href
                                    })
                        name = 'Div: '+str(getClassBasedCSSSelector(driver, selenium_child_divs[0]))
                        if len(hrefs)>3 and len([d for d in data if d['name'] == name])==0:
                            data.append({
                                # 'class': ' '.join(divs['class']),
                                'products': hrefs,
                                'name': name,
                                'hrefs': [d['link'] for d in hrefs]
                            })
                # for size
                first_div_height = selenium_child_divs[0].size['height']
                first_div_width = selenium_child_divs[0].size['width']
                all_same_size = all(div.size['height'] == first_div_height and div.size['width'] == first_div_width for div in selenium_child_divs)
                if all_same_size:
                    hrefs = []
                    if len(selenium_child_divs[0].find_elements(By.CSS_SELECTOR,'img'))==0:
                        continue
                    for w in selenium_child_divs:
                        if len(w.find_elements(By.CSS_SELECTOR,'a[href]'))>0:
                            href = w.find_element(By.CSS_SELECTOR,'a[href]').get_attribute('href')
                            title = getBiggestTitle(w)
                            if len(title.strip())>0:
                                hrefs.append({
                                    'name': title,
                                    'link': href
                                })
                    saveElementScreenShot(driver, getClassBasedCSSSelector(driver, selenium_child_divs[0]), 'product_selector')
                    name = 'Div: '+str(getClassBasedCSSSelector(driver, selenium_child_divs[0]))
                    if len(hrefs)>3 and len([d for d in data if d['name'] == name])==0:
                        data.append({
                            # 'class': ' '.join(divs['class']),
                            'products': hrefs,
                            'name': name,
                            'hrefs': [d['link'] for d in hrefs]
                        })
                    index = index + 1
        # to_delete = []
        # for d in data:
        #     products = []
        #     for h in d['hrefs']:
        #         print(h)
        #         if 'http' not in h:
        #             continue
        #         driver.get(h)
        #         h_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
        #         h_soup = BeautifulSoup(h_res, 'html.parser')
        #         if len(h_soup.select('meta[property="og:title"]'))>0:
        #             try:
        #                 products.append({
        #                     "name": h_soup.select_one('meta[property="og:title"]')['content'],
        #                     "link": driver.current_url,
        #                     # "description": h_soup.select_one('meta[itemprop="description"]')['content'],
        #                     # "image": h_soup.select_one('meta[itemprop="image"]')['content'],
        #                 })
        #             except:
        #                 pass
        #     d['products'] = products
        #     if len(products)==0:
        #         to_delete.append(d)
        # for d in to_delete:
        #     data.remove(d)
        driver.quit()
        return JsonResponse({'data': data})
    
def find_parent_element(product_div, selector, regex):
    while True:
        try:
            product_div.find_element(By.XPATH, "..")
            product_div = product_div.find_element(By.XPATH, "..")
            if len(product_div.find_elements(By.CSS_SELECTOR, selector))>0:
                if regex:
                    if len(re.findall(regex, product_div.text))>0:
                        break
                else:
                    break
        except:
            break
            
    return product_div

def xpath_to_css(xpath):
    # Strip the leading "/" and split by "/"
    parts = xpath.strip('/').split('/')
    
    css_parts = []
    
    for (i, part) in enumerate(parts):
        if len(parts)<i+6:
            if part.startswith('div['):
                # Extract the index (e.g., div[2] -> 2)
                index = re.search(r'\[(\d+)\]', part).group(1)
                css_parts.append(f'div:nth-child({index})')
            elif part == 'html':
                css_parts.append('html')
            elif part == 'body':
                css_parts.append('body')
            elif part == 'img':
                css_parts.append('img')
            else:
                index = re.search(r'\[(\d+)\]', part).group(1)
                part = part.replace(f'[{index}]','')
                # Handle any other tags directly
                css_parts.append(part)
    
    # Join parts with '>'
    return ' > '.join(css_parts)
    # return xpath

def is_element_in_viewport(driver, element):
    # Get the element's position and size
    rect = driver.execute_script(
        "return arguments[0].getBoundingClientRect();", element)
    return rect['top'] >= 0 and rect['left'] >= 0 and rect['bottom'] <= driver.execute_script("return window.innerHeight") and rect['right'] <= driver.execute_script("return window.innerWidth")

class TestInside(APIView):
    def post(self, request, *args, **kwargs):
        url = request.data['url']
        driver = create_browser()
        driver.get(url)
        sleep(10)
        res = {}
        driver.execute_script('''
            document.querySelectorAll('header, footer, iframe').forEach(function(element) {
                element.remove();
            });
        ''')
        biggest_img = None
        for divs in driver.find_elements(By.CSS_SELECTOR,'img[src]'):
            if biggest_img is None or (divs.size['height'] > biggest_img.size['height'] and divs.size['width'] > biggest_img.size['width']):
                if (divs.get_attribute('src') == 'https://whiteangeljo.com/wp-content/uploads/2024/10/2000_66aa2dc4ca82d-400x400.webp'):
                    rect = driver.execute_script("return arguments[0].getBoundingClientRect();", divs)
                    height = driver.execute_script("return window.innerHeight;", divs)
                    width = driver.execute_script("return window.innerWidth", divs)
                    # print(rect)
                    # print(height)
                    # print(width)
                    # print(is_element_in_viewport(driver, divs))
                    # print(divs.size['height'])
                    # print(divs.size['width'])
                if is_element_in_viewport(driver, divs):
                    print(divs.size['height'])
                    print(divs.size['width'])
                    biggest_img = divs
        
        product_div = biggest_img
        res['img'] = {
            'img': biggest_img.get_attribute('src'),
            'xpath': xpath_to_css(getAbsoluteXPath(driver, biggest_img))
        }
        res['imgs'] = {}
        while True:
            product_div = product_div.find_element(By.XPATH, "..")
            if len(product_div.find_elements(By.CSS_SELECTOR, 'img'))>1:
                for img in product_div.find_elements(By.CSS_SELECTOR, 'img'):
                    if img.size['height'] == biggest_img.size['height'] and img.get_attribute('src') != biggest_img.get_attribute('src'):
                        res['imgs']['xpath'] = xpath_to_css(getAbsoluteXPath(driver, img))
                    if img.size['height'] != biggest_img.size['height']:
                        res['imgs']['thumb_xpath'] = xpath_to_css(getAbsoluteXPath(driver, product_div.find_element(By.CSS_SELECTOR, 'img')))
            if len(product_div.find_elements(By.CSS_SELECTOR, 'h1, h2, h3, h4, h5, h6'))>0:
                break
        
        res['title'] = {
            'title': product_div.find_element(By.CSS_SELECTOR, 'h1, h2, h3, h4, h5, h6').text, 
            'xpath':  xpath_to_css(getAbsoluteXPath(driver, product_div.find_element(By.CSS_SELECTOR, 'h1, h2, h3, h4, h5, h6')))
        }
        description_div = biggest_img
        while True:
            try:
                found = False
                description_div = description_div.find_element(By.XPATH,'..')
                for divs in description_div.find_elements(By.CSS_SELECTOR,'div'):
                    # if description_div is None or len(divs.find_elements(By.XPATH, './h1 | ./h2 | ./h3 | ./h4 | ./h5 | ./h6 | ./p | ./span')) > len(description_div.find_elements(By.XPATH, './h1 | ./h2 | ./h3 | ./h4 | ./h5 | ./h6 | ./p | ./span')):
                    if len(divs.find_elements(By.XPATH, './h1 | ./h2 | ./h3 | ./h4 | ./h5 | ./h6 | ./p | ./span'))>3 and len(divs.find_element(By.XPATH, './h1 | ./h2 | ./h3 | ./h4 | ./h5 | ./h6 | ./p | ./span').text.strip())>10:
                        description_div = divs
                        found = True
                        break
                if found:
                    break
            except:
                description_div = biggest_img
                break
            
        res['description'] = {}
        try:
            description_div = getSameFontSize(product_div)
            res['description'] = {
                'description': description_div.text,
                'xpath': xpath_to_css(getAbsoluteXPath(driver, description_div))
            }
        except Exception as e:
            print(e)
        prices = []
        price_div = find_parent_element(biggest_img, 'p, span, bdi, del', r"(JOD|JD)?\s?\d{1,3}\.\d{1,3}\s?(JOD|JD)?")
        for price in price_div.find_elements(By.CSS_SELECTOR, 'p, span, bdi, del'):
            if len(re.findall(r"(JOD|JD)?\s?\d{1,3}\.\d{1,3}\s?(JOD|JD)?", price.text))>0:
                font_weight = price.value_of_css_property('font-weight')        
                if font_weight == 'normal':
                    font_weight_value = 400
                elif font_weight == 'bold':
                    font_weight_value = 700
                else:
                    font_weight_value = float(font_weight)
                font_size = float(price.value_of_css_property('font-size').replace('px',''))
                prices.append({
                    'element': price,
                    'size': font_weight_value,
                    'font_size': font_size,
                })
        biggest_size = 0
        res['price'] = {}
        for p in prices:
            if p['font_size'] > biggest_size:
                biggest_size = p['font_size']
                res['price'] = {
                    'price': p['element'].text,
                    'xpath': xpath_to_css(getAbsoluteXPath(driver, p['element']))
                }
            
        print(res)
        # saveElementScreenShot(driver, xpath_to_css(getAbsoluteXPath(driver, product_div)), 'product_div')
        # saveElementScreenShot(driver, res['img'].get('xpath', 'html'), 'img')
        # saveElementScreenShot(driver, res['imgs'].get('xpath', 'html'), 'imgs')
        # saveElementScreenShot(driver, res['imgs'].get('thumb_xpath', 'html'), 'thumb_imgs')
        # saveElementScreenShot(driver, res['title'].get('xpath', 'html'), 'title')
        # saveElementScreenShot(driver, res['description'].get('xpath', 'html'), 'description')
        # saveElementScreenShot(driver, res['price'].get('xpath', 'html'), 'price')
        driver.quit()
        return JsonResponse({'data': res})

def checkIfItClose(value, values):
    res = []
    for n in values:
        if str(value).strip() == str(n).strip():
            res.append(True)
        else:
            res.append(False)
    return res 

def getSameFontSize(driver):
    res = None
    for divs in driver.find_elements(By.CSS_SELECTOR,'div'):
        # child_divs = divs.find_elements(By.XPATH, './div')
        elements = divs.find_elements(By.XPATH, "./*[text()]")
        font_weights = []
        for element in elements:
            font_weight = element.value_of_css_property('font-weight')        
            if font_weight == 'normal':
                font_weight_value = 400
            elif font_weight == 'bold':
                font_weight_value = 700
            else:
                font_weight_value = int(font_weight)
            font_weights.append(font_weight_value)

        # all_same_size = all(checkIfItClose(font_weights[0], font_weights))
        if len(font_weights)>0:
            for font_weight in font_weights:
                sizes = checkIfItClose(font_weight, font_weights)
                true_count = sizes.count(True)
                percentage_true = (true_count / len(sizes)) * 100
                if percentage_true > 70 and len(font_weights)>3:
                    res = divs
                    break

    return res
        
def getBiggestFont(driver):
    # Find all elements with text on the page (You can refine this to specific tags or classes)
    elements = driver.find_elements(By.XPATH, "//*[text()]")

    # Initialize variables to store the element with the maximum font-weight
    max_font_weight = 0
    max_font_weight_element = None

    # Loop through elements and check their font-weight
    for element in elements:
        font_weight = element.value_of_css_property('font-weight')
        
        # Convert font-weight to a numeric value for comparison (it could be 'normal', 'bold', or numeric)
        if font_weight == 'normal':
            font_weight_value = 400
        elif font_weight == 'bold':
            font_weight_value = 700
        else:
            font_weight_value = int(font_weight)
        
        # Check if this element has the largest font weight
        if font_weight_value > max_font_weight:
            max_font_weight = font_weight_value
            max_font_weight_element = element

    # Output the text of the element with the largest font weight
    if max_font_weight_element:
        print(f"Element with the largest font weight: {max_font_weight_element.text}")
    else:
        print("No element found with a defined font weight.")
        
class TemuScrapView(APIView):
    def post(self, request, *args, **kwargs):
        driver = uc.Chrome(use_subprocess=False)
        url = request.data['url']
        driver.get(url)
        sleep(1)
        data = []
        errors = []
        hrefs = []
        until_visible(driver, '.js-goods-list > div > div a')
        elements = driver.find_elements(By.CSS_SELECTOR, ".js-goods-list > div > div")
        for e in elements:
            if len(e.find_elements(By.CSS_SELECTOR, "a"))>0:
                e.click()
                sleep(5)
                try:
                    driver.switch_to.window(driver.window_handles[1])
                    until_visible_with_xpath(driver, "//div[contains(@aria-label, 'reviews from Jordan')]")
                    href_res = driver.find_element(By.CSS_SELECTOR, 'html').get_attribute('outerHTML')
                    soup = BeautifulSoup(href_res, 'html.parser')
                    if len(soup.select("div[aria-label*='reviews from Jordan'] span")) == 0:
                        return
                    title = soup.select_one("meta[name*='title']")['content']
                    desc = soup.select_one("meta[name*='description']")['content']
                    in_stock = soup.select_one("div[aria-label*='reviews from Jordan'] span").get_text(strip=True).replace('(','').replace(')','').strip()

                    product = {
                        "Name": title,
                        "Description": desc,
                        "Sells": in_stock,
                        "reference_link": driver.current_url,
                    }
                    data.append(product)
                except Exception as e:
                    traceback.print_exc()
                    errors.append({
                        "url": driver.current_url,
                    })
                finally:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
        data.sort(key=lambda x: int(x['Sells']), reverse=True)
        df = pd.DataFrame(data)
        df.to_excel('excel/'+request.data['db_category']+'_products.xlsx', index=False)
        driver.quit()
        return JsonResponse({})

class ChangeText(APIView):
    def post(self, request, *args, **kwargs):
        chrome_options = Options()
        # chrome_options.page_load_strategy = 'eager'
        chrome_options.add_argument('--blink-settings=imagesEnabled=false')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--disable-web-security')
        chrome_options.add_argument('--allow-running-insecure-content')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-setuid-sandbox')
        chrome_options.add_argument('--disable-features=NetworkService')
        chrome_options.add_argument('--disable-features=VizDisplayCompositor')
        chrome_options.add_argument('--disable-features=IsolateOrigins')
        chrome_options.add_argument('--disable-features=AutofillCreditCardSignin')
        # chrome_options.add_argument('--headless')
        # chrome_options.add_argument('--headless')
        driver = Chrome(options=chrome_options)
        driver.maximize_window()
        driver.get('https://www.scribbr.com/paraphrasing-tool/')
        until_visible(driver, '#QuillBotPphrIframe')
        iframe = driver.find_element(By.CSS_SELECTOR, "#QuillBotPphrIframe")
        driver.get(iframe.get_attribute('src'))
        dataframe1 = pd.read_excel('excel/'+request.data['id']+'_products.xlsx')
        products = []
        for index, row in dataframe1.iterrows():
            data = row.to_dict()
            if data['English Description'] != ' ':
                new_desc = change_text(driver, translate(remove_emoji(data['English Description']), dest='en'))
                ar_new_desc = translate(new_desc)
                products.append({
                    "Arabic Name": data['Arabic Name'],
                    "English Name": data['English Name'],
                    "Arabic Description": ar_new_desc if len(ar_new_desc)>0 else data['Arabic Description'],
                    "English Description": new_desc if len(new_desc)>0 else data['English Description'],
                    "Category Id": data['Category Id'],
                    "Arabic Brand": data['Arabic Brand'],
                    "English Brand": data['English Brand'],
                    "Unit Price": data['Unit Price'],
                    "Discount Type": data['Discount Type'],
                    "Discount": data['Discount'],
                    "Unit": data['Unit'],
                    "Current Stock": data['Current Stock'],
                    "Main Image URL": data['Main Image URL'],
                    "Photos URLs": data['Photos URLs'],
                    "Video Youtube URL": data['Video Youtube URL'],
                    "English Meta Tags": data['English Meta Tags'],
                    "Arabic Meta Tags": data['Arabic Meta Tags'],
                    "features": data['features'],
                    "features_ar": data['features_ar'],
                    "wholesale": data['wholesale'],
                    "reference_link": data['reference_link'],
                })
            else: 
                products.append({
                    "Arabic Name": data['Arabic Name'],
                    "English Name": data['English Name'],
                    "Arabic Description": data['Arabic Description'],
                    "English Description": data['English Description'],
                    "Category Id": data['Category Id'],
                    "Arabic Brand": data['Arabic Brand'],
                    "English Brand": data['English Brand'],
                    "Unit Price": data['Unit Price'],
                    "Discount Type": data['Discount Type'],
                    "Discount": data['Discount'],
                    "Unit": data['Unit'],
                    "Current Stock": data['Current Stock'],
                    "Main Image URL": data['Main Image URL'],
                    "Photos URLs": data['Photos URLs'],
                    "Video Youtube URL": data['Video Youtube URL'],
                    "English Meta Tags": data['English Meta Tags'],
                    "Arabic Meta Tags": data['Arabic Meta Tags'],
                    "features": data['features'],
                    "features_ar": data['features_ar'],
                    "wholesale": data['wholesale'],
                    "reference_link": data['reference_link'],
                })

        df = pd.DataFrame(products)
        df.to_excel('excel/new_'+request.data['id']+'_products.xlsx', index=False)
        driver.quit()
        return JsonResponse({})

def remove_classes(driver):
    driver.execute_script("document.querySelectorAll('multistep-form-section > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    driver.execute_script("document.querySelectorAll('multistep-form-section > * > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    driver.execute_script("document.querySelectorAll('multistep-form-section > * > * > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    driver.execute_script("document.querySelectorAll('multistep-form-section > * > * > * > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    driver.execute_script("document.querySelectorAll('multistep-form-section > * > * > * > * > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    driver.execute_script("document.querySelectorAll('multistep-form-section > * > * > * > * > * > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    driver.execute_script("document.querySelectorAll('multistep-form-section > * > * > * > * > * > * > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    driver.execute_script("document.querySelectorAll('multistep-form-section > * > * > * > * > * > * > * > *').forEach((e)=>{e.classList.remove(...e.classList)});")
    sleep(2)

generate_blog_lock = threading.Lock()
class GenerateBlog(APIView):
    def post(self, request, *args, **kwargs):
        blog = Blogs()
        blog.name = request.data['headline']
        blog.status = 'waiting'
        blog.save()
        generate_blog_lock.acquire()
        options = Options()
        options.add_experimental_option('detach', True)
        options.add_argument("--headless") 
        options.add_argument("--no-sandbox") 
        options.add_argument("--disable-dev-shm-usage") 
        options.add_argument("--disable-notifications")
        # options.add_argument("--disable-gpu")
        # options.add_argument("--remote-debugging-port=9222") 
        # options.add_argument(f"crash-dumps-dir={os.path.expanduser('~/tmp/Crashpad')}")
        # options.headless = True

        # Create an instance of Chrome WebDriver
        # driver = webdriver.Firefox(options=options)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        # driver = create_browser()
        try:
            blog.status = 'in progress'
            blog.save()
            driver.save_screenshot('test.png')
            
            # Open the webpage
            driver.get('https://katteb.com/ar/sign-in/')
            driver.maximize_window()
            driver.save_screenshot('test.png')
            email_element = driver.find_element(By.ID, 'username')
            email_element.send_keys("icnnobar@gmail.com")
            password_element = driver.find_element(By.ID, 'password')
            password_element.send_keys("Icn@nobar123")
            driver.save_screenshot('test.png')
            login_button = driver.find_element(By.CSS_SELECTOR, 'button.validation-submit-btn')
            login_button.click()
            driver.save_screenshot('test.png')
            wait = WebDriverWait(driver, 15)
            wait.until(EC.url_contains('/dashboard/'))
            driver.save_screenshot('test.png')
            headline = request.data['headline']
            driver.get('https://katteb.com/ar/dashboard/generate-full-article/')
            driver.save_screenshot('test.png')
            until_visible_click(driver, 'multistep-form-body-field:nth-child(1)')
            driver.save_screenshot('test.png')
            until_visible_send_keys(driver, 'multistep-form-body-field:nth-child(1) input', headline)
            driver.save_screenshot('test.png')
            until_visible_click(driver, '.-step-excerpt')
            sleep(2)
            driver.save_screenshot('test.png')
            # driver.execute_script("document.querySelectorAll('.-sendmessage-qactions').forEach(e => e.remove());")
            # sleep(2)
            # driver.save_screenshot('test.png')
            # driver.execute_script("""
            #     var button = document.createElement('multistep-form-next');
            #     button.className = 'next';
            #     button.textContent = 'test';
            #     var documentToAdd = document.querySelector("multistep-form-section[data-step='1']");
            #     if (documentToAdd) {
            #         documentToAdd.appendChild(button);
            #         console.log('Button added successfully');
            #     } else {
            #         console.log('Target element not found');
            #         document.body.appendChild(button);
            #     }
            # """)
            # driver.save_screenshot('test.png')
            sleep(2)
            # until_visible_click(driver, 'div[data-tooltip="مصدر المحتوى"]')
            # sleep(2)
            # remove_classes()
            # remove_classes(driver)
            driver.execute_script("""
                arguments[0].scrollIntoView({ block: 'center', inline: 'nearest' });
            """, driver.find_element(By.CSS_SELECTOR, "multistep-form-section[data-step='1'] multistep-form-next"))
            driver.save_screenshot('test.png')
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "multistep-form-section[data-step='1'] multistep-form-next"))
            ).click()
            # until_visible_click(driver, "multistep-form-section[data-step='1'] multistep-form-next")
            sleep(10)
            # remove_classes(driver)
            driver.execute_script("""
                arguments[0].scrollIntoView({ block: 'center', inline: 'nearest' });
            """, driver.find_element(By.CSS_SELECTOR, "multistep-form-section[data-step='2'] multistep-form-body-inner > div:last-child > div:last-child"))
            driver.save_screenshot('test.png')
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "multistep-form-section[data-step='2'] multistep-form-body-inner > div:last-child > div:last-child"))
            ).click()
            # until_visible_click(driver, "multistep-form-section[data-step='2'] multistep-form-body-inner > div:last-child > div:last-child")
            sleep(2)
            driver.save_screenshot('test.png')
            show_article = WebDriverWait(driver, 600).until(
                EC.presence_of_element_located((By.LINK_TEXT, 'عرض المقال'))
            )
            driver.save_screenshot('test.png')
            show_article.click()
            driver.save_screenshot('test.png')
            articles_holder = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,
                                                'div.fr-element.fr-view'))
            )
            driver.save_screenshot('test.png')
            ar_output = articles_holder.get_attribute('outerHTML')
            soup = BeautifulSoup(ar_output, 'html.parser')
            headers_components = soup.select('streaming-area')
            res = []
            en_res = []
            for comp in headers_components:
                title = comp.select_one('h2').get_text(strip=True)
                comp.select_one('h2').extract()
                content = comp.get_text(" ", strip=True)
                res.append({
                    'title': title,
                    'description': content,
                })
                en_res.append({
                    'title': translate(title, dest='en'),
                    'description': translate(content, dest='en'),
                })
            keywords = extract_top_keywords(en_res[0]['description'])
            ar_keywords = []
            for keyW in keywords:
                ar_keywords.append(translate(keyW))
                
            data = {'data': json.dumps({
                "category_id": request.data['category'],
                "image": request.data['image'],
                "author": {
                "en": "ICN",
                "sa": "ICN"
                },
                "title": {
                    "en": translate(headline, dest='en'),
                    "sa": headline
                },
                "short_description": {
                    "en": en_res[0]['description'],
                    "sa": res[0]['description']
                },
                'description': {
                    "en": [e['description'] for e in en_res],
                    "sa": [e['description'] for e in res]
                },
                'bookmark': {
                    "en": [e['title'] for e in en_res],
                    "sa": [e['title'] for e in res]
                },
                "meta_description": {
                    "en": en_res[0]['description'],
                    "sa": res[0]['description']
                },
                "meta_keywords": {
                    "en": ','.join(keywords),
                    "sa": ','.join(ar_keywords)
                },
                "meta_title": {
                    "en": translate(headline, dest='en'),
                    "sa": headline
                }
            })}
            try:
                # Send the POST request and wait for the response
                response = requests.post('https://www.icn.com/api/v1/blog/store', data=data)
                
                # Check if the request was successful
                if response.status_code == 200:
                    print('Request was successful!')
                    print('Response:', response.text)  # If the response contains JSON data
                    blog.api_status = 'Request was successful!'
                    blog.save()
                else:
                    print('Request failed with status code:', response.status_code)
                    print('Response:', response.text)
                    blog.api_status = 'Request failed:' + str(response.text)
                    blog.save()
            except requests.exceptions.RequestException as e:
                blog.api_status = 'An error occurred:' + str(e)
                blog.save()
                print('An error occurred:', e)
            blog.status = 'done'
            blog.save()
        except Exception as e:
            print(e)
            blog.status = 'error: ' + str(e)
            blog.save()
        finally:
            driver.quit()
            # Release the lock
            generate_blog_lock.release()
            return JsonResponse({})

class IntegrationTest(APIView):
    def post(self, request, *args, **kwargs):
        driver = uc.Chrome(use_subprocess=False)
        errors = []
        def check_banners(navigate_to='', selector='.banner a'):
            if navigate_to:
                driver.get(navigate_to)
            try:
                until_visible(driver, selector, max_counter=3)
            except:
                pass
            if len(driver.find_elements(By.CSS_SELECTOR, selector))>0:
                hrefs = []
                [hrefs.append(e.get_attribute('href')) for e in driver.find_elements(By.CSS_SELECTOR, selector)]
                for href in hrefs:
                    driver.get(href)
                    try:
                        until_not_visible(driver, 'img[src*="/404.svg"]', counterAmount=2)
                    except Exception as e:
                        traceback.print_exc()
                        errors.append({
                            'page': c,
                            "url": href,
                        })

        def change_lang(href, dest='sa'):
            try:
                driver.get(href)
                until_visible(driver, f'header #change_lang[data-flag={dest}]', max_counter=3)
                until_visible_click(driver, f'header #change_lang[data-flag={dest}]')
                sleep(6)
            except Exception as e:
                print(e)
                pass

        def register():
            driver.get('https://www.icn.com/users/registration')
            until_visible_send_keys(driver, 'input[name="name"]', 'test')
            until_visible_send_keys(driver, 'input[name="phone"]', '111111111')
            until_visible_send_keys(driver, 'input[name="password"]', 'Test$123')
            until_visible_click(driver, '.aiz-square-check')
            until_visible_click(driver, '#submit_reg')
            until_not_visible(driver, 'input[name="name"]')

        def login():
            driver.get('https://www.icn.com/users/login')
            until_visible_send_keys(driver, 'input[name="phone"]', '111111111')
            until_visible_send_keys(driver, 'input[name="password"]', 'Test$123')
            until_visible_click(driver, '#submit_reg')
            until_not_visible(driver, 'input[name="name"]')

        def logout():
            until_visible_click(driver, 'header .dropdown-toggle[aria-expanded="true"]')
            until_visible_click(driver, 'ul.dropdown-menu > li a[href*="logout"]')
            until_not_visible(driver, 'ul.dropdown-menu > li a[href*="logout"]')

        def deleteAccount():
            driver.get('https://www.icn.com/dashboard')
            until_visible_click(driver, '.sidemnenu .aiz-side-nav-item a.confirm-delete-user')
            until_visible_click(driver, '#delete-modal-user .modal-body > a[href="https://www.icn.com/delete/account"]')
            until_not_visible(driver, '#delete-modal-user .modal-body > a[href="https://www.icn.com/delete/account"]')

        try:
            # # home
            # change_lang('https://icn.com/', dest='en')
            # check_banners(selector=".header-row-banner a")
            # check_banners(navigate_to='https://icn.com/')
            # change_lang('https://icn.com/')
            # check_banners(selector=".header-row-banner a")
            # check_banners(navigate_to='https://icn.com/')
            # if len(errors)>0:
            #     df = pd.DataFrame(errors)
            #     df.to_excel(f'excel/home.xlsx', index=False)

            # # category
            # change_lang('https://icn.com/', dest='en')
            # categories = [{'href': e.get_attribute('href'), 'title': e.text} for e in driver.find_elements(By.CSS_SELECTOR, '.nav-categories .swiper-container a:not([href="#"])')]
            # for c in categories:
            #     errors = []
            #     change_lang(c['href'], dest='en')
            #     check_banners()
            #     change_lang(c['href'])
            #     check_banners()
            #     if len(errors)>0:
            #         df = pd.DataFrame(errors)
            #         df.to_excel(f'excel/{c['title']}.xlsx', index=False)
            
            # filter
            change_lang('https://icn.com/filter/category/All', dest='en')
            until_visible(driver, '#accordion > .accordion-item > .accordion-header a')
            elements = [{'href': e.find_element(By.CSS_SELECTOR, '.accordion-header a').get_attribute('href'),'title': e.find_element(By.CSS_SELECTOR, '.accordion-header a').text} for e in driver.find_elements(By.CSS_SELECTOR, '#accordion > .accordion-item')]
            for e in elements:
                # english
                errors = []
                change_lang(e['href'], dest='en')
                check_banners()
                categories = [e.get_attribute('href') for e in driver.find_elements(By.CSS_SELECTOR, '.collapse .accordion-item a')]
                for c in categories:
                    driver.get(c)
                    check_banners()

                # arabic
                change_lang(e['href'])
                check_banners()
                categories = [e.get_attribute('href') for e in driver.find_elements(By.CSS_SELECTOR, '.collapse .accordion-item a')]
                for c in categories:
                    driver.get(c)
                    check_banners()
                    
                if len(errors)>0:
                    df = pd.DataFrame(errors)
                    t = e['title']
                    df.to_excel(f'excel/{t}.xlsx', index=False)
            
        except:
            traceback.print_exc()
        driver.quit()
        return JsonResponse({})

class ImageUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        upload = upload_file(request.FILES['file'], request.data.get('base_id'), request.data.get('table_id'), request.data.get('record_id'))
        if upload:
            return Response({'message': 'Image uploaded successfully', 'status': 200}, status=status.HTTP_201_CREATED) 
        else:
            return Response({'message': 'Image uploaded failed', 'status': 500}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
